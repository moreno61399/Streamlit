import streamlit as st
from openpyxl import reader,load_workbook,Workbook
import io

st.title("FK_KM CHECK")

fkonzept = st.file_uploader("upload FK file", type={"xlsx","csv", "txt"})

wb=Workbook()

if fkonzept is not None:
    wb =load_workbook(fkonzept, read_only=True)
    st.write(wb.sheetnames)
    st.title(wb)
    st.write(wb.active)

wb.close()    
buffer = io.BytesIO()
wb.save(buffer)


st.download_button(
    label="Download Excel worksheet without index",
    data=buffer.getvalue(),
    file_name="fk.xlsx",
)
    
 
