# -*- coding: utf-8 -*-
"""
Created on Mon Feb  6 11:38:03 2023

@author: eduardo.moreno
"""

from openpyxl import load_workbook, Workbook
import pandas as pd

import fk.fk as fk
import km.km as km
import df.df as DF

import io
buffer = io.BytesIO()

import streamlit as st


import sys
import fk.vt.vt as vt



def convert_df(df):
   return df.to_csv(index=False).encode('utf-8')

def convert_df_excel(df):
   return df.to_csv(index=False).encode('utf-8')

def get_dict_vts(wb_fk, Projekt = '', delete_grey = 'no'):
    ''' Final method, the one that you ask for '''    
    #cargamos el archivo
    #wb_fk = get_as_wb(fk_path)
    
    #buscamos el BauRaum del fk introducido
    
# =============================================================================
#     wb=wb_fk
# =============================================================================
    dict_vts = {}
    sheets = wb_fk.sheetnames

    for name in sheets:
       
        try:

            if ('vt' in name.lower() or 'pim' in name.lower() or name.lower().startswith('steu')) and ('PLS' not in name):
               
                ws = wb_fk[name]
                                
                if ws.sheet_properties.tabColor== None or ws.sheet_properties.tabColor.rgb != 'FFFF0000':
                    vt.delete_red_striked_vt(ws)
                    if delete_grey == 'yes':                    
                        vt.delete_grey_vt(wb_fk,ws)
                    
                    df_vt_aux = pd.DataFrame(ws.values)   
                    df_vt = vt.clean_vt(df_vt_aux, name)
                    
                    name = name.strip()
                    dict_vts[name] = df_vt.reset_index(drop=True)    
                
        except Exception as error:
            print('Error cleaning ' + name + '.\n Error: <{}>'.format(error))
            sys.exit()
                    
# =============================================================================
#                 name = name + '_' + BauRaum + '_' + HandDrive
# =============================================================================
       
    return dict_vts



def main_check_FK(Status, wb_fk, df_KM, delete_grey):

    #para evitar que se quede siempre abierta la ventana emergente de tkinter
    #root = Tk()
    #root.withdraw()
    #obtenemos de la KM-Liste los módulos internos y los metemos en una lista
    df_KM_aktuell =  pd.DataFrame()
    
    df_KM.columns = list(map(lambda x: x.strip(),df_KM.columns))
   
    st.write(df_KM.columns)
    df_KM_aktuell = km.filter_by_Status(df_KM, Status)
   
    #y generamos un diccionario VT-df_VT con el df_VT ya limpio de rojo y tachado
    dict_vts = get_dict_vts(wb_fk, '', delete_grey)
    st.write(dict_vts)

    #introducimos una parte para leer desde aquí también las hojas de wires y boms

    # LOAD FK AS WB  
    #wb = fk.get_as_wb(fk_path)
    
    # GET REAL NAMES FROM FK ASSIGNED TO WIRES AND BOMS
    lista_names = fk.get_list_names_wires_and_boms(wb_fk)
    
# =============================================================================
#     df_prueba = pd.read_excel(fk_path,sheet_name = lista_names[0])
# =============================================================================
    
    st.write(lista_names)
    st.write(wb_fk)
    
    # GET THE DATAFRAMES FROM WIRES AND BOMS
    print("Reading WIRES BOMs")
    (df_wire, df_bom), (pos_headers_wire, pos_headers_bom) = fk.get_df_wires_and_boms(wb_fk, lista_names)
    print("Finish Reading WIRES BOMs")
    # Cambiar los nombres de columnas repetidas en Wire
    columns_change = pd.Series(df_wire.columns)
    columns_change = columns_change.fillna(value = 'column_name')
    df_wire.columns = columns_change
    df_wire = DF.unmangleCols2(df_wire)   
   
    # Cambiar los nombres de columnas repetidas en BOM
    columns_change = pd.Series(df_bom.columns)
    columns_change = columns_change.fillna(value = 'column_name')
    df_bom.columns = columns_change
    df_bom = DF.unmangleCols2(df_bom)

    #generamos el report de errores así como el diccionario de vts ya corregido
    dict_Report_final, dict_vts_corregido = fk.get_Report_FK(dict_vts, df_KM_aktuell,df_wire, df_bom)
    
    
    #AQUÍ VIENE LA DESCARGA DEL EXCEL
    
    
    #y lo pasaoms a un excel que se puede enviar a los compañeros de CPE
    contador_errores = 0
    
    for Report_name, df_Report in dict_Report_final.items():
        if len(df_Report) > 0:
            contador_errores = contador_errores + 1
                
    if contador_errores == 0:
        st.write("Check_FK", "Congratulations!\n\nNo errors have been detected!")
    elif contador_errores >= 1:
        
        with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
            for Report_name, df_Report in dict_Report_final.items():
                if len(df_Report)>0:
                    df_Report.to_excel(writer,sheet_name=Report_name,index=False)
        writer.save()
        st.download_button(
        label="Download Excel worksheet without index",
        data=buffer,
        file_name="df1.xlsx",
        mime="application/vnd.ms-excel",)
        if contador_errores ==1:
            st.write("Check_FK", 'Oh no!\n\n' + str(contador_errores) + ' error has been detected.\n\nPlease check file "Report_Check_FK"')
        elif contador_errores > 1:
            st.write("Check_FK", 'Oh no!\n\n' + str(contador_errores) + ' errors have been detected.\n\nPlease check file "Report_Check_FK"')


st.title("FK_KM CHECK")
wb = Workbook()

ruta_fk = st.text_input('Ruta_Fk', 'ruta')
st.write(ruta_fk)
fkonzept = st.file_uploader("upload FK file", type={"xlsx","csv", "txt"})
workbook_xml = BytesIO(fkonzept)
workbook_xml.seek(0)
wb = openpyxl.load_workbook(workbook_xml)


km_liste = st.file_uploader("upload KM file", type={"xlsx","csv", "txt"})


#st.title(fkonzept.name)




df_fk=pd.DataFrame()
df_km=pd.DataFrame()

if fkonzept is not None:
    df_fk = pd.read_excel(fkonzept)
    #wb = load_workbook(ruta_fk,data_only=True)
    st.title(wb)
    st.write(wb.active)
    
if km_liste is not None:
    df_km = pd.read_excel(km_liste,engine='openpyxl')    
    
    
 
#Status = st.text_input("Enter your name", "")
#Status=int(float(Status))

Status=2325
excel=main_check_FK(Status, wb, df_km,"yes")


st.write(df_fk)


#♂st.title(fk)

csv = convert_df(df_fk)

st.download_button(
   "Press to Download your REPORT",
   excel,
   "file.csv",
   "text/csv",
   key='download-csv'
)
