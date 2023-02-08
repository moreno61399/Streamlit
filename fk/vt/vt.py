# -*- coding: utf-8 -*-
"""
Created on Tue Oct 13 09:25:19 2020
**********************************************************************************************************************************
*  The purpouse of this file (python module), is to be called
*  from another one having the clean_fk.get_dict_fk() dictionary
*  in a variable. For example,
*  imagin that we import this file (python module) into another,
*  we could do the following:
*
*  import clean_fk
*  import vt                                       # Could be another path (take it just as an example) 
*
*  my_dict_fk = clean_fk.get_dict_fk()             # Dictionary with key (VT name) and (VT DataFrame) as value
*
*  for vt_name in my_dict_fk:
*            variant_nummerier_dict = vt.get_variants_nummerierungs_dict(my_dict_fk[vt_name]) # vt DataFrame as a parameter
*            #  the same with the other methods...
*
*  Hope you enjoy it :p
************************************************************************************************************************************
"""

########## NECESSARY IMPORT PARENT MODULE ##########################################
import sys
import getpass
import numpy as np
import df.df as df
import km.km as km
import pandas as pd
import Excel.format as form

from tkinter import *
from itertools import product
from tkinter import messagebox
from openpyxl import Workbook



# =============================================================================
# import sys
# sys.path.append('km/')
# =============================================================================



#get a Series with all variants (names) presentin a VT
def get_variants(df_vt):
    variants =  df_vt.columns[15:].to_series().reset_index(drop=True)
    return variants


def get_dict_modul_nummerierung_with_VT(df_vt, vt_name):
    #Para el chequeo de Sumas tambien necesitamos sacar la relación Módulo - Nummerierung
    df_modul_Num_aux = df_vt.iloc[:,[3,2]]
    df_modul_Num_aux.loc[:,'Nummerierung'] = df_modul_Num_aux.loc[:,'Nummerierung'].astype(str) + '_' + vt_name
    
    #convertimos el df MV intern - Nummerierung en diccionario...
    dict_modul_Num_aux = df.convert_df_to_dict(df_modul_Num_aux)
    
    return dict_modul_Num_aux


def get_df_variant_moduls(df_vt, variants = None):

    try:

        df_vt_basis = df_vt.loc[:,['Nummerierung','MV intern']].reset_index(drop=True)
        df_variants = df_vt.iloc[0:, 15:].reset_index(drop=True)
        
        # df_nummerierungs + df_variants DataFrame
        df_mod_var = pd.concat([df_vt_basis, df_variants], axis=1, ignore_index=False)
        
        # Variant list
        if variants == None:
            variants = list(df_variants.columns)
    
        df_variant_moduls = pd.DataFrame()                                                     # Key: variant, Value: Nummerierungs

        for variant in variants:
            # Filter for X nummerierung
            df_variant = df_mod_var[(df_mod_var[variant]=='X') | (df_mod_var[variant]=='x')]
            
            list_nums = df_variant['Nummerierung'].drop_duplicates().tolist()
            
            dict_num_moduls =dict()
            for num in list_nums:
                group = df_variant[df_variant['Nummerierung']==num]['MV intern'].tolist()
                dict_num_moduls[num] = group

            new_combination_list = get_all_combinations(dict_num_moduls)
            
            df_variant_combinations = pd.DataFrame(new_combination_list)
            df_variant_combinations.insert(0,'Variant',variant)
            df_variant_moduls = pd.concat([df_variant_moduls,df_variant_combinations],ignore_index=True)
                            
        return df_variant_moduls
    
    except KeyError as e:
        sys.exit(str(e) + ' not present in this VT')
    except Exception as err:
        sys.exit(str(err) + ': This VT seems not to be OK. Possible duplicated variant name')



#la variable more_than_one hace referencia a si en una sola estructura puede venir más de un módulo de la misma familia
def get_all_combinations(dict_num_moduls,more_than_one=None):
    
    #faltaría introducir aqúi la opción de incluir únicamente aquellas combinaciones que no permitan más de un módulo por familia
      
    list_nums  =list(dict_num_moduls.keys())
    
    list_of_lists =[]
    for num in list_nums:
        list_of_lists.append(dict_num_moduls[num])
        
    combination_list = list(product(*list_of_lists))
    list_of_combinations = []
            
    for elements in combination_list:
        set_elements= set(elements)
        list_of_combinations.append(set_elements)   

    return list_of_combinations



def get_dict_variants_nummerierungs(df_vt):
    ''' 
        Taking the df of a VT as a parameter,
        returns a dictionary with Variant as a key and
        the corresponding nummerierungs as value.
    '''
    try:   
        df_nummerierungs = df_vt['Nummerierung'].astype(int).to_frame().reset_index(drop=True)
        df_variants = df_vt.iloc[0:, 15:].reset_index(drop=True)
        
        # df_nummerierungs + df_variants DataFrame
        df_num_var = pd.concat([df_nummerierungs, df_variants], axis=1, ignore_index=False)
        
        # Variant list
        variants = list(df_variants.columns)
    
        dict_variant = {}                                                           # Key: variant, Value: Nummerierungs
    
        for variant in variants:
                                                     # Filter for X nummerierung
                df_variant = df_num_var[(df_num_var[variant]=='X') | (df_num_var[variant]=='x')]                       # in every variant
                list_nummerierungs = df_variant['Nummerierung'].unique().tolist()
                dict_variant[variant] = list_nummerierungs
        
        return dict_variant
    
    except KeyError as e:
        print(str(e) + ' not present in this VT ' + get_vt_name(df_vt))
        messagebox.showerror('Error running vt.get_dict_variants_nummerierungs',str(e) + ' not present in VT ' + get_vt_name(df_vt))

    except Exception as err:
        print('VT ' + get_vt_name(df_vt) + ' seems not to be OK. Possible duplicated variant name\n\n' + str(err))
        messagebox.showerror('Error running vt.get_dict_variants_nummerierungs', 'VT ' + get_vt_name(df_vt) + ' seems not to be OK. Possible duplicated variant name\n\n' + str(err))


def get_dict_variants_nummerierungs_with_VT(df_vt,vt_name):
    ''' 
        Taking the df of a VT as a parameter,
        returns a dictionary with Variant as a key and
        the corresponding nummerierungs as value.
    '''

    df_nummerierungs = df_vt['Nummerierung'].astype(str) + '_' + vt_name
    df_nummerierungs = df_nummerierungs.to_frame().reset_index(drop=True)

    df_variants = df_vt.iloc[0:, 15:].reset_index(drop=True)
    
    # df_nummerierungs + df_variants DataFrame
    df_num_var = pd.concat([df_nummerierungs, df_variants], axis=1, ignore_index=False)
    df_num_var.reset_index(inplace = True)

    # Variant list
    variants = list(df_variants.columns)

    dict_variant = {}                                                           # Key: variant, Value: Nummerierungs

    for variant in variants:                                                    # Filter for X nummerierung
        df_variant = df_num_var[(df_num_var[variant]=='X') | (df_num_var[variant]=='x')]                       # in every variant
        list_nummerierungs = df_variant['Nummerierung'].unique().tolist()
        dict_variant[variant] = list_nummerierungs


    return dict_variant

# =============================================================================
# def VT_exceptions (Projekt,df_differences):
#     
#     if Projekt == 'J1':
#         dict = {'COCKPIT_RL_VT02':'COCKPIT_LL_VT02','COCKPIT_RL_VT05':'COCKPIT_LL_VT05'}
#     
#     df_differences = pd.read_excel('I:/Ingenieria/0_DptoIng/00_KM_FK/J1_PAG_Taycan/0_Estatus_Activos/2149/Jit-Calls Check/2022/Reception on KW04/Report_Indices_y_Sumas_OK_only_Cockpit_VT05.xlsx',sheet_name = 'Report_Sumas')
#         
# =============================================================================


def check_int_index(df_VT):
    try:
#creo un nuevo DataFrame llamdo df_of_comparar con las columnas 'MV intern','int. Index'              
        df_of_comparar = df_VT.iloc[:,[1,3,5]]
#en la columna MV intern me quedo con su ultimo caracter                
        df_of_comparar['index MV intern'] = df_of_comparar['MV intern'].str[-1:]
#borro todos los NAN, espacios en blanco por que sino no me los imprime en el resultado 
        df_of_comparar = df_of_comparar.dropna(how='all')
#convertimos ambas columnas a string para que no genere diferencias por tipos diferentes
        df_of_comparar['index MV intern']= df_of_comparar['index MV intern'].astype(str)
        df_of_comparar['int. Index']= df_of_comparar['int. Index'].astype(str)
#creo un nuevo DataFrame llamado solución el cual compara los elementos de las dos tablas (MV intern,int. Index) y se rellena de los valores que son diferentes                
        Report_int_index = pd.DataFrame(df_of_comparar[df_of_comparar['index MV intern'] !=  df_of_comparar['int. Index']])
#Añado una nueva columna llamada VT al DataFrame 'solución' en dicha columna se añaden los nombres de las hojas que hayan cumplido las condiciones puestas hasta ahora                  
        Report_int_index['VT'] = get_vt_name(df_VT)

#Para terminar, se borran todos los NAN que se hayan podido colar y se reemplaza el DataFrame con NAN por el sin NAN               
        Report_int_index.dropna(how='all',inplace=True)
        
        return Report_int_index
        
    except:
       print("Problems by function check_int_index")

#Esta nueva función se encarga de asegurar que siempre que los Num/IBG, también estén siempre marcados con "x" o "-" al mismo tiempo
def check_Variants_vs_NumIBG(df_VT):
    
    df_Num_IBG = df_VT.iloc[:,[2,7]] 
    df_Variants = df_VT.iloc[:,15:]
    df_new = pd.concat([df_Num_IBG,df_Variants],axis=1)
    df_new.insert(0,'Num_IBG',df_new.iloc[:,0].astype(str) + '_' + df_new.iloc[:,1])
    df_new.drop(columns=df_new.columns[1:3],inplace=True)
    
    list_variants = get_variants(df_VT).to_list()
    
    list_Num_IBG = df_new['Num_IBG'].unique().tolist()
    
    dict_errors = dict()
    
    for Num_IBG in list_Num_IBG:
        df = df_new[df_new['Num_IBG']==Num_IBG]
        if len(df)>1:
            lista_variants_erroneas = []
            for variant in list_variants:
                serie_to_check = df[variant].unique()
                if len(serie_to_check)>1:
                    lista_variants_erroneas.append(variant)
            if len(lista_variants_erroneas)!=0:
                dict_errors[Num_IBG] = lista_variants_erroneas
    
    if len(dict_errors.keys())!=0:
        Num_IBG_with_error = list(dict_errors.keys())
        variants_with_error = list(dict_errors.values())[0]
        variants_with_error.insert(0,'Num_IBG')
    
        Report_Inconsistence_Num_vs_Variant= df_new[df_new['Num_IBG'].isin(Num_IBG_with_error)].loc[:,variants_with_error]
    else:
        Report_Inconsistence_Num_vs_Variant = pd.DataFrame()
        
    return Report_Inconsistence_Num_vs_Variant


    
def duplicated_nummerierung_values(df_VT):
    
 try:
    df=df_VT.iloc[:,[1,2,4,7]]
    df.loc[:,'Nummerierung'] = df.loc[:,'Nummerierung'].astype(str).str.strip()
    df.loc[:,'Infobaugruppe NEU / aktuell'] = df.loc[:,'Infobaugruppe NEU / aktuell'].astype(str).str.strip()
    
    duplicated_rows = df[df.duplicated(subset=['Nummerierung'])]        # Procesamos el <df> para quedarnos solo con los duplicados...
    duplicated_nummeriers = duplicated_rows['Nummerierung'].to_list()
    index_of_duplicated = [index for index, value in df['Nummerierung'].items() if value in duplicated_nummeriers]
    
    dpdf = df.iloc[index_of_duplicated].sort_values('Nummerierung').reset_index(drop=True)  # DataFrame con valores duplicados
    #primero en un sentido...
    nummerierung = dpdf['Nummerierung'].unique().tolist()               
    output_df=pd.DataFrame()
    for num in nummerierung:
        df_aux = dpdf[dpdf['Nummerierung']==num]                        # Esta DF auxiliar contiene el dpdf subdividido por cada valor nummerierung.
        df_aux2 = df_aux.drop_duplicates(subset='Infobaugruppe NEU / aktuell')                  # Solo nos interesan los nummerier repetidos pero con valor <neu> DISTINTO.
        if len(df_aux2) > 1:
            output_df = pd.concat([output_df, df_aux])
    #y ahora en el otro..
    new_duplicated_rows = df[df.duplicated(subset=['Infobaugruppe NEU / aktuell'])]        # Procesamos el <df> para quedarnos solo con los duplicados...
    new_duplicated_nummeriers = new_duplicated_rows['Infobaugruppe NEU / aktuell'].to_list()
    new_index_of_duplicated = [index for index, value in df['Infobaugruppe NEU / aktuell'].items() if value in new_duplicated_nummeriers]
    
    new_dpdf = df.iloc[new_index_of_duplicated].sort_values('Infobaugruppe NEU / aktuell').reset_index(drop=True)  # DataFrame con valores duplicados
    
    list_IBG = new_dpdf['Infobaugruppe NEU / aktuell'].unique().tolist()
    for ibg in list_IBG:
        df_aux_ibg = new_dpdf[new_dpdf['Infobaugruppe NEU / aktuell']==ibg]                        # Esta DF auxiliar contiene el dpdf subdividido por cada valor ibg.
        df_aux_ibg2 = df_aux_ibg.drop_duplicates(subset='Nummerierung')                  # Solo nos interesan los ibg repetidos pero con valor <num> DISTINTO.
        if len(df_aux_ibg2) > 1:
            output_df = pd.concat([output_df, df_aux_ibg])

    if len(output_df)!=0:
        output_df = output_df.drop_duplicates()
        output_df[output_df.columns[0]]=get_vt_name(df_VT)
        return output_df

 except:
       print("Problems by function duplicated_nummerierung_values")

#para obtener el nombre del VT como string

def get_vt_name(df_VT):
    
    #ponemos en primer lugar el caso excepcion de la hoja steuerungsvorgabe
    if str(df_VT.columns[15][0])=='2':
        VT = 'Steuerungsvorgabe'
    elif str(df_VT.columns[15][0])=='9':
        VT = 'VTWWL'
    else:       
        VT_from_variant = str(df_VT.columns[15][6:8])
    
        if (VT_from_variant[0]!=VT_from_variant[1] and VT_from_variant[1]!='0' and VT_from_variant[0]!='0')\
            or VT_from_variant.isdigit():
            VT = VT_from_variant
        elif VT_from_variant[0] == VT_from_variant[1] or VT_from_variant[1] == '0':
            VT = str(VT_from_variant[0])
        elif VT_from_variant[0] == '0':
            VT = str(VT_from_variant[1])
    return VT


#comprueba si hay variantes que estén fuera de formato, diferentes de X o -
def check_invalid_data(df_VT):

    df_Variantes = df_VT.loc[:,df_VT.columns[15:]]
#    df_Variantes = df_Variantes.loc[:,df_Variantes.columns[15:]].replace(to_replace = ' -',value = '-')
#    df_Variantes = df_Variantes.loc[:,df_Variantes.columns[15:]].replace(to_replace = '- ',value = '-')
#    df_Variantes = df_Variantes.loc[:,df_Variantes.columns[15:]].replace(to_replace = ' X',value = 'X')
#    df_Variantes = df_Variantes.loc[:,df_Variantes.columns[15:]].replace(to_replace = 'X ',value = 'X')
    
    Report_invalid_data = pd.DataFrame()

    for variante in df_Variantes.columns:
        print(variante)
        Error = df_Variantes[(df_Variantes[variante].str.strip()!='X') &\
                             (df_Variantes[variante].str.strip()!='-') &\
                             ((df_Variantes[variante].str.strip()!='L'))].loc[:,variante]
        if len(Error)>=1:
            df_Variantes_Error = df_Variantes[variante].to_frame()
            Report_invalid_data = pd.concat([Report_invalid_data,df_Variantes_Error],axis=1)

    if len(Report_invalid_data)>0:
        Report_invalid_data = pd.concat([Report_invalid_data,df.extract_ca(df_VT,df_VT.columns[[1,3]])],axis=1)
        Report_invalid_data = Report_invalid_data.sort_index(axis=1,ascending=False)
    return Report_invalid_data

#comprueba si el nombre de la variante tiene algún error
def check_variant_name(df_VT):
    '''
    Chequea el nombre de las variates:
        - El nombre debe tener 13 caracteres

    Parameters
    ----------
    df_VT : DataFrame 
        Datos del VT.

    Returns
    -------
    df_VT : DataFrame
        Datos del VT con los errores en los nombres corregidos.
    Report_variant_name :   DataFrame
        Reporte de variantes con errores en el nombre.

    '''
    
    VT = get_vt_name(df_VT)
    Variantes = pd.Series(df_VT.columns[15:], name = 'IST')
    df_Variantes = Variantes.to_frame()
    df_Variantes['VT'] = VT
    #para los casos en los que tenemos VT "A", por ejemplo y las variantes debe contener "AA" o "A0"
    if len(VT) < 2:
        
        if VT == str(df_Variantes.loc[0, 'IST'])[6]:
            VT = VT + '0'
        elif VT == str(df_Variantes.loc[0, 'IST'])[7]:
            VT = '0' + VT
        
    if VT == "VTWWL":
        df_report = df_Variantes[(df_Variantes.IST.str[0] != '9')]
    else:
        df_report = df_Variantes[
            ((df_Variantes['IST'].str[3] != '3')| (df_Variantes['IST'].str.len() != 13) |\
            ((df_Variantes['IST'].str[6:8] != VT[-2:])))]
        
        df_report.loc[:,'SOLL'] = df_report.loc[:,'IST'].str[:3] + '3' + df_report.loc[:,'IST'].str[4:6] + VT[-2:] + df_report.loc[:,'IST'].str[8:]
        Report_variant_name = df_report[['VT','IST','SOLL']].reset_index()
        
        for index in range(0,len(Report_variant_name)):
            df_VT = df_VT.rename(columns={Report_variant_name.loc[index,'IST']:Report_variant_name.loc[index,'SOLL']})
            
    return df_VT, Report_variant_name

def check_headers(df_VT):
    
    #establecemos los encabezados oficiales
    Official_VT = pd.read_excel('h:/Groups/CPE/0_DptoIng/6_Instrucciones oficiales/Anweisungen_CPE/CPE-W-074_Darstellung FK/CPE-W-074 Rev.9/CPE-W-074 Rev. 9 Anlage 4 Zuordnungsmatrix (Vorlage) - Attachment 4 Assignment matrix (template).xlsx',engine='openpyxl')
    
    Official_headers = Official_VT.iloc[7,:].dropna().tolist()
    
    Official_headers_reduced = Official_headers[1:6]
    
    VT_headers = df_VT.columns.to_list()
    
    VT_headers_reduced = VT_headers[1:6]
    
    
    #definimos el VT
    VT = get_vt_name(df_VT)
    #recorremos las columnas del VT
    Report_headers = pd.DataFrame(columns=['VT','IST','SOLL'])
    #aquí establecemos los títulos de columnas que queremos checkear
    #columnas_clave = [1,2,3,4,5,7]
    for element in range(0,len(Official_headers_reduced)):
        #comprobamos si es o no
        if VT_headers_reduced[element]==None or VT_headers_reduced[element]!=VT_headers_reduced[element] or (VT_headers_reduced[element].strip() != Official_headers_reduced[element].strip() and Official_headers_reduced[element].strip() != 'frei verfugbar'):
            new_df = pd.DataFrame([[VT,VT_headers_reduced[element],Official_headers_reduced[element]]], columns=['VT','IST','SOLL'])
            Report_headers = pd.concat([Report_headers,new_df],ignore_index=True)
            df_VT = df_VT.rename(columns={VT_headers_reduced[element]:Official_headers_reduced[element]})
    
    return df_VT,Report_headers

def check_SEBN_moduls_KM_Liste(df_VT,df_KM_aktuell):
    try:
        df_modulos_vt= df.extract_ca(df_VT,['MV intern', 'MV extern'])
        df_KM_aktuell = df.extract_ca(df_KM_aktuell, ['Einsatz','Entfall','MV extern','MV intern'])
        df_KM_aktuell =df_KM_aktuell.rename(columns={'MV intern':'MV intern in KM-Liste'})
        df_merged = df_modulos_vt.merge(df_KM_aktuell,on='MV extern',how = 'left')
        Report_moduls_KM_Liste = df_merged[df_merged['MV intern']!=df_merged['MV intern in KM-Liste']]
        Report_moduls_KM_Liste['VT']=get_vt_name(df_VT)
        Report_moduls_KM_Liste=Report_moduls_KM_Liste[['VT','MV extern','MV intern','MV intern in KM-Liste']]
        Report_moduls_KM_Liste = Report_moduls_KM_Liste.drop_duplicates()
    except Exception as e:
        print(str(e),': Error detected on VT ' + get_vt_name(df_VT))
        messagebox.showerror('Error running vt.get_dict_variants_nummerierungs', 'VT ' + get_vt_name(df_vt) + ' seems not to be OK. Possible duplicated variant name\n\n' + str(err))

        
    return Report_moduls_KM_Liste

'''cambiar esta función'''
def check_IBG_KM_Liste(df_VT, df_KM_aktuell):

    #obtenemos el nombre del VT
    vt_name = get_vt_name(df_VT)
    #sacamos por un lado la relacion módulo-IBG del FK del VT a analizar
    
    df_FK = df.extract_ca(df_VT, ['MV intern','Infobaugruppe NEU / aktuell'])
    df_FK['VT']=vt_name
    df_FK = df_FK[['VT','MV intern','Infobaugruppe NEU / aktuell']]
    
    #y por otro la relación módulo-IBG de la KM-Liste del VT a analizar
    df_moduls_aktuell_with_IBG = km.get_moduls_aktuell_with_IBG(df_KM_aktuell)
    vt_name_KM = 'VT' + vt_name
    
    if vt_name_KM not in df_moduls_aktuell_with_IBG.columns and \
       len(df_moduls_aktuell_with_IBG.columns[df_moduls_aktuell_with_IBG.columns.str.contains(vt_name)]) > 0:       
        vt_name_KM = df_moduls_aktuell_with_IBG.columns[df_moduls_aktuell_with_IBG.columns.str.contains(vt_name)].values[0]
    
    Report_IBG_final = pd.DataFrame()
    if vt_name_KM in df_moduls_aktuell_with_IBG.columns:
        df_IBG = df_moduls_aktuell_with_IBG.loc[:,['MV intern',vt_name_KM]]
        df_IBG = df_IBG.rename(columns={vt_name_KM:'IBG in KM-Liste'})
        
        df_merge = df_FK.merge(df_IBG,on='MV intern',how='left')
        Report_IBG = df_merge[df_merge.iloc[:,2]!= df_merge.iloc[:,3]]
        
        #añadimos estas lineas adicionalemente para el caso de TR, donde en cada celda de la KM-Liste puede haber más de un IBG
        for fila in range(0,len(Report_IBG)):
    
            IBG_FK = str(Report_IBG.iloc[fila,2]).strip()
            IBG_KM = str(Report_IBG.iloc[fila,3]).strip()
            if IBG_KM!=IBG_KM or IBG_FK not in IBG_KM:            
                Report_IBG_final = pd.concat([Report_IBG_final,Report_IBG.iloc[fila].to_frame().T])
            
    return Report_IBG_final

def get_real_values(ws):

    pos_headers = 9
    #buscamos la max fila y columna que contienen datos de esa hoja
    maximum_row = max ((c.row for c in ws['E'] if c.value is not None))
    maximum_column = max ((c.column for c in ws[pos_headers] if c.value is not None))
    
    #generamos un nuevo workbook con una nueva hoja para dejar ahí los datos que nos interesan
    new_book = Workbook()
    new_ws = new_book[new_book.sheetnames[0]]
    
    for row in ws.iter_rows(max_row=maximum_row, max_col=maximum_column):
        new_ws.append((cell.value for cell in row))

    return new_ws

def delete_red_striked_vt(ws):
    
    '''
        Removes all the rows/columns with red and striked values
    '''
    
# =============================================================================
#     from openpyxl import load_workbook
#     file_path = 'C:/Users/jesus.roldan/Desktop/D1021771-6-F_Konzept_PO513_PAGJ1_FGRLL_TAB052652_KW47_21_ZD11062021_6.xlsm'
#     wb = load_workbook(file_path)
#     ws = wb['VTB']
# =============================================================================
    # ws = get_real_values(ws)
    hoja = ws.title
    
    maximum_row = max ((c.row for c in ws['E'] if c.value is not None))
    ws.delete_rows(maximum_row + 1, ws.max_row)
    maximum_column = max ((c.column for c in ws[9] if c.value is not None))
    ws.delete_cols(maximum_column + 1, ws.max_column)
    
    #Primero recorremos filas basándonos en la columna E (Módulo interno)
    cont = 0
    for cell in ws['E']:
        
        fila = cell.row
        
        if cell.value!=None:
            if cell.font.strike == True:
                if cell.font.color == None or (cell.font.color and cell.font.color.rgb !='FFFF0000'):   #el texto está tachado y tiene color rojo
                    print('Cuidado! Fila ' + str(fila + cont) + ' en hoja ' + str(hoja) + ' está tachada pero no en rojo y ha sido eliminada!')
                ws.delete_rows(fila)
                cont = cont+1
    
    #Después recorremos columnas basándonos en la fila 6 (Número de variante)
    contador = 0
    for celda in ws[6]:
  
        columna = celda.column
        
        if celda.value!=None:
            if celda.font.strike == True:
                if celda.font.color == None or (celda.font.color and celda.font.color.rgb !='FFFF0000'):  #el texto está tachado y tiene color rojo
                    print('Cuidado! Columna ' + str(columna + contador) + ' de hoja ' + str(hoja) + ' está tachada pero no en rojo y ha sido eliminada!')
                ws.delete_cols(columna) 
                contador = contador + 1
    

        
    return ws

def delete_grey_vt(wb_fk,ws):
    
    '''
        Removes all the rows/columns with grey values
    '''
    
# =============================================================================
#     from openpyxl import load_workbook
#     file_path = 'C:/Users/jesus.roldan/Desktop/D1021771-6-F_Konzept_PO513_PAGJ1_FGRLL_TAB052652_KW47_21_ZD11062021_6.xlsm'
#     wb = load_workbook(file_path)
#     ws = wb['VTB']
# =============================================================================
    
    hoja = ws.title

    #Primero recorremos filas basándonos en la columna E (Módulo interno)
    cont = 0
    for cell in ws['E']:
        fila = cell.row
        if cell.value!=None:
            if cell.fill.start_color.tint:
                theme_x = cell.fill.start_color.theme
                tint_x = cell.fill.start_color.tint
                x = str(form.theme_and_tint_to_rgb(wb_fk, theme_x, tint_x))
            else:
                x = str(cell.fill.start_color.rgb)[:6]
                
            if x[0:2] == x[2:4] and x[2:4] == x[4:] and x != '000000' and x != 'FFFFFF' or x == 'FFDDDD':
                print('Ojo! Fila ' + str(fila + cont) + ' en hoja ' + str(hoja) + ' en gris ha sido borrada')
                ws.delete_rows(fila)
                cont = cont+1
                
                
    #Después recorremos columnas basándonos en la fila 6 (Número de variante)
    contador = 0
    for celda in ws[6]:
  
        columna = celda.column
        if celda.value!=None:
            if celda.fill.start_color.tint:
                theme_y = celda.fill.start_color.theme
                tint_y = celda.fill.start_color.tint
                y = str(form.theme_and_tint_to_rgb(wb_fk, theme_y, tint_y))
            else:
                y = str(celda.fill.start_color.rgb)[:6]
                
            if y[0:2] == y[2:4] and y[2:4] == y[4:] and y != '000000' and y != 'FFFFFF' or y == 'FFDDDD' :
                print('Ojo! Columna ' + str(columna + contador) + ' en hoja ' + str(hoja) + ' en gris ha sido borrada')
                ws.delete_cols(columna)
                contador = contador+1




# =============================================================================
# def delete_red_striked_vt(ws):
#     
#     '''
#         Removes all the rows/columns with red and striked values
#     '''
#     
#     #Primero recorremos filas basándonos en la columna E (Módulo interno)
#     for cell in ws['E']:
#         if cell.font.strike:
#             ws.delete_rows(cell.row)
#         #Después recorremos columnas basándonos en la fila 6 (Número de variante)
#         for celda in ws[6]:
#             if celda.font.strike:  #el texto está tachado y tiene color rojo
#                 ws.delete_cols(celda.column)
#                 
#     df_vt = pd.DataFrame(ws.values)   
#                              # From openpyxl woorkbook to pandas df_vt
# 
#     return df_vt
# =============================================================================





def clean_vt(df_vt, name):
    ''' 
        Deletes any possible mistake, like random text in wrong cells
        and replaces the variant tokens cells with the corresponding
        v1, v2, v3, v4, ... cell to be able to take all the data as a perfect rectangular table.  
    '''    

    df_vt = df_vt.drop(df_vt.columns[0], axis=1)                                         # Drops the first column(always empty)
    # Deletes all full None columns and we don't get Nontype + str error doing the ''.join... below
    #df_vt = df_vt.dropna(axis=1, how='all')
    if name == "VTWWL":
        df_variants_tokens = df_vt.iloc[[4, 5, 6, 7], 15:].dropna(axis=0,how = 'all')
        df_variants_tokens = df_variants_tokens.dropna(axis=1,how = 'all')
    else:
        df_variants_tokens = df_vt.iloc[[4, 5, 6, 7], 15:].dropna(axis=1)
        # A df_vt with the variant tokens cell,
                                                                               # always starts at col 'Q' and row '5' . . .
    df_vt=df_vt.iloc[:,:15+int(len(df_variants_tokens.iloc[0]))]
    
    variants_list = []
                                                                                # We concat all the tokens from every variant
    for i in range(len(df_variants_tokens.iloc[0])):    
                                                                                # and we put them on the list.
        variant_as_list = df_variants_tokens[df_variants_tokens.columns[i]].tolist()
        variant_name = ''
        for variant_as in variant_as_list[::-1]:
            variant_name = variant_name + str(variant_as)
# =============================================================================
#         reversed_variant = ''
# =============================================================================
        
        variants_list.append(variant_name)
# =============================================================================
#         for token in reversed_variant_list:
#             
#             reversed_variant += str(token)
#             
#         variants_list.append(reverse_variant(reversed_variant))
# 
# =============================================================================
    


    df_vt.iloc[8, 15:] = [variant for variant in variants_list]                    # Replace the v1, v2, ... 
    df_vt = df_vt.iloc[8:,:]
    df_vt=df_vt.dropna(how='all',axis=0)                                                        # with the corresponding variant
    df_vt=df_vt.dropna(subset=[df_vt.columns[2]])
    
    
    #esta función se usaba pero si la ejecutamos nos dejamos cosas sin checkear en el main_check_FK
    #df_vt = remove_bad_rows(df_vt)
    
    ####### DROPS THE MISTAKES NEXT TO VARIANT TABLE
    #df_vt.iloc[:, 15:] = df_vt.iloc[:, 15:].dropna(axis=1)
    #df_vt = df_vt.dropna(axis=1,how='all')
    #######
    
    # Replaces the first row with the columns headers
    new_header = df_vt.iloc[0]                                                     # Grab the firs row for the header
    df_vt = df_vt[1:]                                                                 # take the data less header
    df_vt.columns = new_header                                                     # set the header row as the df_vt header
    
    ###### _0 mv extern error solution ########
    df_vt.iloc[:,4] = get_parsed_mvextern(df_vt)
    ###########################################

    # Sustituimos los posibles espacios de las variantes como " -", "- ", "x ", " x"
    a1 = 'x'
    df_vt.iloc[:,15:] = df_vt.iloc[:,15:].replace(to_replace = a1,value = a1.upper())
    df_vt.iloc[:,15:] = df_vt.iloc[:,15:].replace(to_replace = ' ' + a1.upper(),value = a1.upper())
    df_vt.iloc[:,15:] = df_vt.iloc[:,15:].replace(to_replace = a1.upper() + ' ',value = a1.upper())
    df_vt.iloc[:,15:] = df_vt.iloc[:,15:].replace(to_replace = ' -',value = '-')
    df_vt.iloc[:,15:] = df_vt.iloc[:,15:].replace(to_replace = '- ',value = '-')
    
    return df_vt                                                 # Removes the None values
                                                                                # May be better to put this at the begining...

def get_parsed_mvextern(df_vt):
    '''
        Checks for '_0' characters at the end of MV extern value,
        and deletes them. (Just the '_0' part) 

    Parameters
    ----------
    mvextern_column : SERIE
        MV extern column as a SERIE

    Returns
    -------
    mvextern_parsed : SERIE
        MV extern column without any '_0' value
        at the end.

    '''
    mvextern_column = df_vt.iloc[:,4]
    mvextern_parsed = mvextern_column.map(lambda value: value[:-2] if value[-2:] == '_0' else value)
    
    return mvextern_parsed


def remove_bad_rows(df_vt):
    ''' 
        Removes the rows with human mistakes
        if the cell value is None or it doesn't have
        the mv_intern format, we replace it with 'eRr0r'.
        Finally we delete all the rows with some eRr0r value.
    '''
    
    f = lambda x: x if x is not None and (len(str(x)) == 13) and x.startswith('1') else 'eRrOr' 
    # Para mayor precisión, añadir más condiciones
    df_vt.iloc[1:,3] = df_vt.iloc[1:,3].map(f)
    df_vt = df_vt[df_vt[4] != 'eRrOr']                                                    # Mv intern is col 4 
    
    
    return df_vt


# =============================================================================
# def reverse_variant(reversed_variant):
#     ''' 
#         From 'C1001LUAA1J13' to '1L13LUAA001C' 
#         This is neccesary because when we take the variant tokens,
#         we do from top to bottom.
#     '''
#     try:
#         variant = reversed_variant[-4:] + reversed_variant[-8:-4] + reversed_variant[2:5] + reversed_variant[0:2]
#     except Exception as error:
#         print('Error with a variant format (it may be the variant size).\n Error <{}>'.format(error))
#         sys.exit()
#         
#     return variant
# 
# =============================================================================


