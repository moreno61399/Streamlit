# -*- coding: utf-8 -*-
"""
Created on Wed Jan 13 12:01:08 2021

@author: jesus.roldan
"""

from openpyxl import load_workbook, Workbook
from PIL import ImageColor
from openpyxl.styles import PatternFill, Font
from tkinter import filedialog
from tkinter import messagebox
from copy import copy
#import streamlit as st

import re
import os
import sys
import pandas as pd
import fk.vt.vt as vt
import df.df as df
import fk.wires.wires as wires
import fk.boms.boms as boms
import dict.dict as dic
import Excel.format as form

def get_Sesam_Doc(fk_path):

    file_name = os.path.basename(fk_path)
    
    Position = file_name.find('D')
    
    if Position ==-1:
        sesam_doc = 'DXXXXXX'
    else:
        sesam_doc = file_name[Position:Position+8]

# =============================================================================
#     sesam_version = file_name.split('-')[1].split('-')[0]
#     
#     sesam_doc_version = sesam_doc + '-' + sesam_version
# =============================================================================
    
    return sesam_doc



##Ojo al parámetro data_only ya que se lee solo los datos del wb
def get_as_wb(file_path):
    wb_fk = load_workbook(file_path,data_only=True)
    return wb_fk

def get_list_names_wires_and_boms(wb):
    
    list_names=[]
    OPTIONS_WIRES = ['wire', 'WIRE','Wirelist','wirelist','WIRELIST','WIRES','wires']
    OPTIONS_BOMS = ['bom', 'BOM','bomlist','BOMLIST','BOMS','boms', 'boom']        

    for sheet in wb.sheetnames:
        sheet_l = sheet.lower()
        if any(x in sheet_l for x in OPTIONS_WIRES):
           list_names.append(sheet)
    
    for sheet in wb.sheetnames:
        sheet_l = sheet.lower()
        if any(x in sheet_l for x in OPTIONS_BOMS):
           list_names.append(sheet)

    return list_names


def get_colors_before(wb, lista_names):
    '''
    Función encargada de obtener los colores de una lista de hojas excel (solo 
    tiene en cuenta los colores rojo y gris). Para ello recorre cada una de las
    celdas si el color de la fuente es rojo adiciona al valor de la celda el string 
    '_red' y si el color de la fuente es gris adiciona al valor de la celda el 
    string '_grey'.

    Parameters
    ----------
    wb : Workbook
        Contiene la información del excel.
    lista_names : lista de String
        Lista de los nombres de las hojas del excel a analizar.

    Returns
    -------
    wb : Workbook
        Contiene la información del excel con los colores añadidos.

    '''    
    try:
     
        for sheet in lista_names:
            ws  = wb[sheet]
            RED    = 'FFFF0000' 
            YELLOW = 'FFFFFF00'
            BLUE   = 'FF0000FF'
            GREEN  = 'FF00FF00'
            #GREY   = (166,166,166)
     
            for row in ws.iter_rows(ws.min_row, ws.max_row):
                for cell in row:
                    
                    # Chequeamos que la celda no esté vacía
                    if cell.value == None:
                       cell.value = '' 
                    
                    if cell.fill != None:
                        if cell.fill.start_color.tint:
                            theme_x = cell.fill.start_color.theme
                            tint_x = cell.fill.start_color.tint
                            x = form.theme_and_tint_to_rgb(wb, theme_x, tint_x)
                        else:
                            x = str(cell.fill.start_color.rgb)[:6]
                       
                        if x[0:2] == x[2:4] and x[2:4] == x[4:] and x != '000000'and\
                           x != 'FFFFFF' or x == 'FFDDDD':
                           cell.value = str(cell.value) + '_GREY' 
                           
                    # Chequeamos que la celda tenga color
                    if cell.font.color == None:
                       continue
                        
                    if len(str(cell.font.color.rgb)) == 8:       # Hex value (RRGGBB)
                       
                        color = str(cell.font.color.rgb)
                        
                        if color == RED:
                           cell.value = str(cell.value) + '_red'
                        
                        if str(cell.fill.fgColor.rgb) == YELLOW:
                           cell.value = str(cell.value) + '_yellow' 
                           
                        if str(cell.fill.fgColor.rgb) == GREEN:
                           cell.value = str(cell.value) + '_green'    
                        
                        if color == BLUE:
                           cell.value = str(cell.value) + '_blue'
                        
                        # Color gris
                        x = str(cell.font.color.rgb)
                        if x[2:4] == x[4:6] and x[4:6] == x[6:]:
                           cell.value = str(cell.value) + '_grey'  
                    else:
                        if cell.font.color.theme == 6 or (cell.font.color.tint != 0.0 and \
                           (cell.font.color.theme == 0 or cell.font.color.theme == 1)):
                           cell.value =  str(cell.value) + '_grey'
    except:
        print("El valor ha de ser 'BOM' o 'WIRE'")
    
    return wb

def get_real_values(wb, sheet_name,key_col):

    sheet = wb[sheet_name]
    
    pos_headers = 0
    for cell in sheet['A']:
        if cell.value == 'AEM':
            pos_headers = cell.row
            break
    
    #buscamos la max fila y columna que contienen datos de esa hoja
    maximum_row = max ((c.row for c in sheet[key_col] if c.value is not None))
    maximum_column = max ((c.column for c in sheet[pos_headers] if c.value is not None))
    
    #generamos un nuevo workbook con una nueva hoja para dejar ahí los datos que nos interesan
    new_book = Workbook()
    new_ws = new_book[new_book.sheetnames[0]]
    new_ws.title = sheet_name
    
    for row in sheet.iter_rows(max_row=maximum_row, max_col=maximum_column):
        for cell in row:
            new_cell = new_ws.cell(row=cell.row, column=cell.col_idx,
                    value= cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    return new_ws


def get_df_wires_and_boms(wb, lista_names):

    lista_df = []
    lista_pos_headers = []

    for i in range(len(lista_names)):
        print('reading',lista_names[i])
        
        if i ==0:
            new_ws = get_real_values(wb,lista_names[i],'C')
            new_ws = wires.delete_red_striked_wires(new_ws)
            
        elif i==1:
            new_ws = get_real_values(wb,lista_names[i],'B')
            new_ws = boms.delete_red_striked_boms(new_ws)
        
        data = new_ws.values
        
# =============================================================================
#         columns=next(data)[0:]
# =============================================================================
        
        df =pd.DataFrame(data)
        
# =============================================================================
#         df = pd.DataFrame(data, columns=columns)
# =============================================================================

        #df = pd.read_excel('parsed_fk.xlsx',sheet_name=i)
        
        
        pos_headers = 0
        for j in range(df.shape[0]):
            if str(df.iloc[j, 0]) == 'AEM':
                pos_headers = j
                break
            
        headers = df.iloc[pos_headers]
        df_with_headers = df.iloc[pos_headers + 1:]
        df_with_headers.columns = headers
        df_with_headers = df_with_headers.reset_index(drop = True) 
        lista_df.append(df_with_headers)
        lista_pos_headers.append(pos_headers)

    return lista_df, lista_pos_headers

def set_colors_after(wb, sheet_name):
    '''
    Función encargada de poner los colores en una hoja de excel. Recorre cada 
    una de las celdas de la hoja, si la celda contiene el string '_yellow' pone 
    el background de la celda en amarrillo, si contiene '_blue' cambia el color 
    de fuente a azul, si contiene el '_red' cambia el color de fuente a rojo y si
    si contiene el '_grey' cambia el color de fuente a gris.

    Parameters
    ----------
    wb : Workbook
        Contiene la información del excel.
    sheet_name : String
        Nombre de la hoja del excel a analizar.

    Returns
    -------
    None.
    '''
    
    # Obtenemos la hoja a analizar
    ws = wb[sheet_name]
    
    # Color de background amarillo    
    YELLOWFILL = PatternFill(start_color = 'FFFF00', end_color = 'FFFF00', \
                             fill_type   = 'solid')
    # Color de background verde    
    GREENFILL = PatternFill(start_color = '00FF00', end_color = '00FF00', \
                             fill_type   = 'solid')
    # Color de background gris 
    GREYFILL = PatternFill(start_color = 'A6A6A6', end_color = 'A6A6A6', \
                              fill_type   = 'solid')    
        
    # Recorremos todas las celdas
    for row in ws.iter_rows(ws.min_row, ws.max_row):
        con = 0
        for cell in row:
            
            # Poner color de font azul
            if '_blue' in str(cell.value):
                cell.font  = Font(color = '0000FF')
                cell.value = str(cell.value).replace('_blue', '')
                
            # Poner color de font rojo
            if '_red'  in str(cell.value):
                cell.font  = Font(color = 'FF0000', strike = True)
                cell.value = str(cell.value).replace('_red', '')
                
            # Poner color de font gris
            if '_grey' in str(cell.value):
                cell.font  = Font(color = 'A6A6A6')
                cell.value = str(cell.value).replace('_grey', '')
            
            # Color verde de background, indica que se debe verificar manualmente su información
            if '_green' in str(cell.value):
                cell.fill  =  GREENFILL
                cell.value = str(cell.value).replace('_green', '')
            
            # Poner color de fondo gris
            if '_GREY' in str(cell.value):
                cell.fill  = GREYFILL
                cell.value = str(cell.value).replace('_GREY', '')
            
            # Poner color de background amarillo
            if '_yellow' in str(cell.value):
                cell.fill  = YELLOWFILL
                cell.value = str(cell.value).replace('_yellow', '')
            
            if con > 0 and str(cell.value).isnumeric() == True:
               cell.value = int(cell.value) 
            
            num_format = re.compile("[0-9]*\.[0-9]+$")
            isfloat = re.match(num_format, str(cell.value))
            if isfloat:
               cell.value = str(cell.value).replace('.', ',')
            
            con = con + 1
            
            
def get_BauRaum(fk_path):
    options_INRA = ['INRA','IR', 'INNENRAUM','INNEN']
    options_MORA = ['MORA','MR','MOTORRAUM']
    options_COCKPIT = ['COCKPIT']
    options_FGSR = ['FGSR','FGR','FGST','FAHRGASTRAUM']
    options_Fahrwerk = ['Fahrwerk','FAHRWERK','FARHWERK','FAWE']
    options_Tueren = ['Tueren','turen','TUEREN','TUREN']
    options_Vorderwagen = ['VORDERWAGEN', 'VOWA','VDW']
    
    if any(x in fk_path.upper() for x in options_INRA):
        BauRaum = 'INRA'
    elif any(x in fk_path.upper() for x in options_MORA):
        BauRaum = 'MORA'
    elif any(x in fk_path.upper() for x in options_COCKPIT):
        BauRaum = 'COCKPIT'
    elif any(x in fk_path.upper() for x in options_FGSR):
        BauRaum = 'FGST'
    elif any(x in fk_path.upper() for x in options_Fahrwerk):
        BauRaum = 'FAHRW'
    elif any(x in fk_path.upper() for x in options_Tueren):
        BauRaum = 'TUEREN'
    elif any(x in fk_path.upper() for x in options_Vorderwagen):
        BauRaum = 'VWAGEN'
    else:
        print('Caution! BauRaum for',fk_path.split('/')[-1],'not found')
        BauRaum = 'not found'
    
    return BauRaum

def get_HandDrive(fk_path):
    '''
        Search the directory resources for LOL and LOR
        files.
        
        Returns a list with [lol_files, lor_files]
        
    '''
    
    OPTIONS_LL = ['LL','L0L','LOL']
    OPTIONS_RL = ['RL','L0R','LOR','LR', 'ROL']
    
    if (any(x in fk_path for x in OPTIONS_LL) and \
        any(x in fk_path for x in OPTIONS_RL)):
        HandDrive = 'LL/RL'
    
    elif any(x in fk_path for x in OPTIONS_LL):
        HandDrive = 'LL'

    elif any(x in fk_path for x in OPTIONS_RL):
        HandDrive = 'RL'
    else:
        print('Caution! Hand Drive for',fk_path.split('/')[-1],'not found')
        HandDrive = 'not_found'
            
    return HandDrive

# =============================================================================
# fk_path = 'C:/Users/jesus.roldan/Desktop/ejemplo tr/prueba.xlsx'
# delete_grey = 'yes'
# =============================================================================

def get_dict_vts(fk_path, Projekt = '', delete_grey = 'no'):
    ''' Final method, the one that you ask for '''    
    #cargamos el archivo
    wb_fk = get_as_wb(fk_path)
    
    #buscamos el BauRaum del fk introducido
    
# =============================================================================
#     wb=wb_fk
# =============================================================================
    dict_vts = {}
    sheets = wb_fk.sheetnames

    for name in sheets:
       
        try:

            if ('vt' in name.lower() or 'pim' in name.lower() or name.lower().startswith('steu')) and ('PLS' not in name):
               
                ws = wb_fk[name]
                                
                if ws.sheet_properties.tabColor== None or ws.sheet_properties.tabColor.rgb != 'FFFF0000':
                    vt.delete_red_striked_vt(ws)
                    if delete_grey == 'yes':                    
                        vt.delete_grey_vt(wb_fk,ws)
                    
                    df_vt_aux = pd.DataFrame(ws.values)   
                    df_vt = vt.clean_vt(df_vt_aux, name)
                    
                    name = name.strip()
                    dict_vts[name] = df_vt.reset_index(drop=True)    
                
        except Exception as error:
            print('Error cleaning ' + name + '.\n Error: <{}>'.format(error))
            sys.exit()
                    
# =============================================================================
#                 name = name + '_' + BauRaum + '_' + HandDrive
# =============================================================================
       
    return dict_vts

def get_all_variants(dict_vts):

    #generamos una serie de objetos auxiliares
    VT_variants = pd.Series(dtype='object')
    
    for vt_name, df_vt in dict_vts.items():

        #Para el checkeo de índices vamos sacando las variantes presentes en cada VT...
        VT_variants_aux =vt.get_variants(df_vt)
        #...y las vamos concatenando en una serie que contenga todas las variantes válidas
        VT_variants=pd.concat([VT_variants,VT_variants_aux],ignore_index=True)
        
    return VT_variants

#depending on FK configuration, this function can take too much time
def get_df_variant_moduls(dict_vts_df_vts_all_FKs):

    dict_VT_variant_modul = {}
    
    for vt_name, df_vt in dict_vts_df_vts_all_FKs.items():
            
        dict_variant_modul_aux = vt.get_df_variant_moduls(df_vt)
        
        dict_VT_variant_modul.update(dict_variant_modul_aux)

    df_variant_modul =  pd.DataFrame.from_dict(dict_VT_variant_modul,orient='index').reset_index()

    return df_variant_modul

def get_dict_variant_Num(dict_vts):
    #generamos una serie de objetos auxiliares

    dict_VT_variant_Num = dict()
    
    for vt_name, df_vt in dict_vts.items():
        try:
            #Para el chequeo de Sumas vamos sacando los Nummerierungs asignaddos a cada VT...
            dict_VT_variant_Num_aux = vt.get_dict_variants_nummerierungs_with_VT(df_vt, vt_name)
            #...y los vamos concatenando en un diccionario que contenga todas las variantes posibles con sus Nummerierungs
            dict_VT_variant_Num.update(dict_VT_variant_Num_aux)
        except Exception as e:
            print('Some problem occured on ' + vt_name)
            messagebox.showinfo("Error running get_dict_variant_Num", 'Some problem occured on ' + vt_name +'\n\n' + str(e))

            
    return dict_VT_variant_Num


def get_dict_modul_Num(dict_vts):

    dict_modul_Num=dict()
    
    for vt_name, df_vt in dict_vts.items():

        dict_modul_Num_aux =vt.get_dict_modul_nummerierung_with_VT(df_vt, vt_name)
        
        #y vamos concatenando todos los diccionarios en uno único que nos proporcione esa relacion Modulo - Nummerierung
        dict_modul_Num = dic.dict_update(dict_modul_Num,dict_modul_Num_aux)

    return dict_modul_Num



def get_Report_FK(dict_vts, df_KM_aktuell, df_wire, df_bom):

    #dataframes auxiliares para ir concatenando la información de los diferentes VTs
    Report_int_index_final = pd.DataFrame()
    Report_nummerierung_final = pd.DataFrame()
    Report_invalid_data_final = pd.DataFrame()
    Report_variant_name_final = pd.DataFrame()
    Report_headers_final = pd.DataFrame()
    Report_SEBN_moduls_KM_Liste_final = pd.DataFrame()
    Report_IBG_final = pd.DataFrame()
    Report_dup_variant = pd.DataFrame()
    Report_dup_variant_name = pd.DataFrame()
    Report_missing_vts = pd.DataFrame()
    Report_Inconsistence_Num_vs_Variant = pd.DataFrame()
    
    aux_dup_variant_name = pd.DataFrame()
    
    dict_vts_corregido = {}
    for VT in dict_vts:
        
        #value=len(dict_vts)
        #st.progress(value, text="Analyzing FK and KM")
        
        #POR AHORA NO HACEMOS CHECKEO DE LA HOJA STEUERUNGSVORGABE
        if not VT.lower().startswith('st'):
            Report_variant_name = pd.DataFrame()
            df_VT = dict_vts[VT]
            #ahora se trata solo de aplicarlas diferentes funciones creadas para el VT
            
            #función 1: Comprobar que los títulos de las columnas clave de los VTs siguen el standard. Reportar y corregir
            df_VT, Report_headers = vt.check_headers(df_VT)
            Report_headers_final = pd.concat([Report_headers_final, Report_headers],ignore_index=True)
            
            #función 2: Comprobar nombre variante como por ejemplo que el cuarto dígito es siempre "3". Reportar y corregir.
            if VT != 'VTWWL':
                df_VT, Report_variant_name = vt.check_variant_name(df_VT)
            Report_variant_name_final = pd.concat([Report_variant_name_final, Report_variant_name],ignore_index=True)
            
            # función 3: Variantes repetidas en nombre. Solo reportar
            variants = pd.DataFrame()
            variants['VT'] = [VT] * len(df_VT.columns[15:])
            variants['Variants'] = df_VT.columns[15:]
            dup_variants = variants[variants['Variants'].duplicated()]
            Report_dup_variant_name = pd.concat([Report_dup_variant_name, dup_variants], ignore_index = True)
            
            # función 4: Variantes repetidas en contenido. Solo reportar            
            dict_modules = vt.get_dict_variants_nummerierungs_with_VT(df_VT, VT)
            df_modules = pd.DataFrame.from_dict(dict_modules, orient = 'index')
            dup = df_modules.duplicated(keep = False)
            dup_variant = df_modules[dup]
            dup_variant.insert(0, 'Variants', df_modules.index[dup])
            dup_variant = dup_variant.reset_index(drop = True) 
            if len(dup_variant) > 0: 
               dup_variant = dup_variant.sort_values(by = list(dup_variant.columns[1:]))
            
            Report_dup_variant = pd.concat([Report_dup_variant, dup_variant], ignore_index = True)
            
            #función 5: Comprobar celdas vacías en las variantes. Solo reportar
            if len(dup_variants) == 0:
                Report_invalid_data = vt.check_invalid_data(df_VT)
                Report_invalid_data_final = pd.concat([Report_invalid_data_final, Report_invalid_data], ignore_index = True)
                  
            #funcion 6: comprobar índices internos del MV intern vs. la columna int.index. Solo Reportar
            Report_int_index = vt.check_int_index(df_VT)
            Report_int_index_final = pd.concat([Report_int_index_final, Report_int_index],ignore_index=True)
            
            #función 7: comprobar si los Nummerierungs e IBG concuerdan. Solo Reportar
            Report_nummerierung = vt.duplicated_nummerierung_values(df_VT)
            Report_nummerierung_final = pd.concat([Report_nummerierung_final, Report_nummerierung],ignore_index=True)
            
            #función 8: comprobar que siempre que Nummerierungs/IBG son iguales, siempre están tachados para asegurar. Solo reportar
            Report_Inconsistence_Num_vs_Variant_aux = vt.check_Variants_vs_NumIBG(df_VT)
            Report_Inconsistence_Num_vs_Variant_aux.insert(0,'VT',VT)
            Report_Inconsistence_Num_vs_Variant = pd.concat([Report_Inconsistence_Num_vs_Variant, Report_Inconsistence_Num_vs_Variant_aux],ignore_index=True)
                        
            if df_KM_aktuell.empty == False:
                
                #función 9: comprobar MV intern con el de la KM-Liste. Solo reportar
                Report_SEBN_moduls_KM_Liste = vt.check_SEBN_moduls_KM_Liste(df_VT, df_KM_aktuell)
                Report_SEBN_moduls_KM_Liste_final = pd.concat([Report_SEBN_moduls_KM_Liste_final, Report_SEBN_moduls_KM_Liste],ignore_index=True)
                
                #función 10: comprobar IBG con el de la KM-Liste. Solo reportar
                Report_IBG = vt.check_IBG_KM_Liste(df_VT, df_KM_aktuell)
                Report_IBG_final = pd.concat([Report_IBG_final,Report_IBG],ignore_index=True)
        
                # Función 11: comprobar que aparezcan todos los VT del FK en la KM-Liste
                if VT not in df_KM_aktuell.columns:
                    aux = pd.DataFrame(data={'VT Name' : [VT]})
                    Report_missing_vts = pd.concat([Report_missing_vts, aux], ignore_index=True)
    
    
            #como se han hecho algunas correcciones, generamos un nuevo diccionario del FK con los VTs corregidos
            dict_vts_corregido[VT] = df_VT


    '''A continuación introducimos todas aquellas funciones de checkeo que no se realizan en los VTs'''
    #función 10: comprobar Baugruppe vs. Verbauort en wires. Solo reportar
    Report_Baugruppe_wires_final = wires.check_Baugruppe(df_wire)
    Report_Baugruppe_wires_final.insert(0,'Sheet','Wires')
    
    #función 11: comprobar Baugruppe vs. Verbauort en wires. Solo reportar
    
    Report_Baugruppe_boms_final = boms.check_Baugruppe(df_bom)
    Report_Baugruppe_boms_final.insert(0,'Sheet','Boms')





    #y por otro lado generamos el diccionario del reporte de errores
    dict_Report_final = {'Report_headers' : Report_headers_final,\
                         'Report_int_index' : Report_int_index_final,\
                         'Report_Nummerierung' : Report_nummerierung_final,\
                         'Report_invalid_data' : Report_invalid_data_final,\
                         'Report_variant_name' : Report_variant_name_final,\
                         'Report_IBG_KM_Liste' : Report_IBG_final,\
                         'Report_Baugruppe_wires' : Report_Baugruppe_wires_final,\
                         'Report_Baugruppe_boms' : Report_Baugruppe_boms_final,\
                         'Report_SEBN_moduls_KM_Liste' : Report_SEBN_moduls_KM_Liste_final,\
                         'Report_Dupl_variants_name' : Report_dup_variant_name,\
                         'Report_Dupl_variants_content' : Report_dup_variant,\
                         'Report_Missing VT in KM-Liste' : Report_missing_vts,\
                         'Report_Num_vs_Variant': Report_Inconsistence_Num_vs_Variant}
        
    return dict_Report_final, dict_vts_corregido

def load_fk_file(path=None):
    '''
    Permite al usuario obtener el path del file donde se encuentra el fk, si se 
    le pasa por defecto el path, devuelve el mismo dado como entrada.
    
    Parameters
    ----------
    path : string, default=None
    
    Attributes
    ----------

    Returns
    -------
    path : String
        Cadena de caracteres que representa el path del file 
    '''
    
    root = Tk()
    root.withdraw()
    
    
    if path is None:
        try:
            path = filedialog.askopenfilename(title="Select the fk file",\
                                              filetypes=[("Excel files","*.xls?")])
                                              
        except Exception as error:
            print("Error loading the fk file: <{}>".format(error))
    
    return path


def load_folder(fk_folder_path=None):
    '''
    Permite al usuario obtener el directorio donde se encuantran los fk.
    Busca en el directorio los file que sean fk y los clasifica en 'LL' o 'RL',
    empleando 'options_LL' y 'options_RL' respectivamente, para ello verifica si 
    estas opciones aparecen en la cadena de caracteres que representa el path del fk.     
     
    Parameters
    ----------
    path : string, default=None
    
    Attributes
    ----------
    options_LL : array de String
        Contine las opciones que son válidas para LL
    options_RL: array de String
        Contine las opciones que son válidas para RL
  
    Returns
    -------
    lol_files : array
        lista de fk que su guía es LL             
    
    lor_files : array
        lista de fk que su guía es RL  
    '''
    
    root = Tk()
    root.withdraw()
    
    if fk_folder_path is None:
    
        try:
            fk_folder_path = filedialog.askdirectory(title="Select the folder where the fk files are located")
        except Exception as error:
            print("Error loading the fk folder: <{}>".format(error))
    
    return fk_folder_path

def skip_open_files(list_files):
    '''
    Esta función recibe una lista de direcciones de documentos y desecha los
    que estén abiertos. Si pasas un str, lo convierte a lista y luego lo vuelve 
    a convertir a str     
     
    Parameters
    ----------
    list_files : list/str
    
    Returns
    -------
    skip_open_list : list/str
    '''
    convert = False
    if isinstance(list_files, str):
        print('only str')
        list_files = [list_files]
        convert = True
        
    skip_open_list = [doc for doc in list_files if "~$" not in doc]
    trash_list = [doc for doc in list_files if "~$"  in doc]
    
    if trash_list:
        print("OPEN documents have been skipped:")
        print(trash_list)
    
    if convert:
        skip_open_list = "".join(skip_open_list)
    
    return skip_open_list

def filter_by_HandDrive(fk_folder_path, guide='ALL'):
    

    list_fk_files = os.listdir(fk_folder_path)
    
    #Filtramos los docs abiertos
    list_fk_files = skip_open_files(list_fk_files)
    
    list_fk_files_filtered = []
    
    for fk_name in list_fk_files:
        
        if guide =='ALL':
            list_fk_files_filtered.append(os.path.join(fk_folder_path, fk_name))

        elif guide == 'LL':
            if any(x in fk_name for x in ['LL','L0L','LOL', 'ALL']):
                list_fk_files_filtered.append(os.path.join(fk_folder_path, fk_name))
        elif guide == 'RL':
            if any(x in fk_name for x in ['RL','L0R','LOR','LR', 'ROL', 'ALL']):
                list_fk_files_filtered.append(os.path.join(fk_folder_path, fk_name))

    return list_fk_files_filtered


# =============================================================================
# list_fk_files_filtered = ['C:/Users/jesus.roldan/Desktop/ejemplo tr/prueba.xlsx']
# Projekt = 'TR'
# delete_grey = 'yes'
# =============================================================================



def get_dict_fk_dict_vts(list_fk_files_filtered, Projekt = '', delete_grey = 'no'):
    
    ''' 
        @Parameter:
         - list with file_paths of the fk files
            
        Returns a list with all the lor files as a dict, key: vt value df_vt 
    '''
    dict_fk_dict_vts = {}
    
    if list_fk_files_filtered:

        for fk_path in list_fk_files_filtered:

            Sesam_doc = get_Sesam_Doc(fk_path)
            BauRaum = get_BauRaum(fk_path)    
            HandDrive = get_HandDrive(fk_path)

            #y generamos un diccionario VT-df_VT con el df_VT ya limpio de rojo y tachado
            dict_vts = get_dict_vts(fk_path, Projekt, delete_grey)
            
            #siempre válido (por ahora)
            if Projekt == 'BG' and BauRaum == 'FAHRW':
                var1, var2 = 'PIM_03' , 'VTS_08'
                dict_vts[var1] = vt_merger(dict_vts[var1],dict_vts[var2],[var1, var2]) 
                dict_vts.pop(var2,None)
            #solo válido a partir de 0-Serie
            if Projekt == 'BG' and BauRaum == 'INRA' and HandDrive == 'LL':
                var1, var2 = 'VT01' , 'VT11'
                dict_vts[var1] = vt_merger(dict_vts[var1],dict_vts[var2],[var1, var2]) 
                dict_vts.pop(var2,None)
                
            #siempre valido en 0-serie
            if Projekt == 'BG' and BauRaum == 'VWAGEN':
                var1, var2 = 'VT09' , 'VT10'
                dict_vts[var1] = vt_merger(dict_vts[var1],dict_vts[var2],[var1, var2]) 
                dict_vts.pop(var2,None)

            #solo válido en 0-Serie_2. En 0-Serie_1 NO EXISTIA EL VT13
            if Projekt == 'BG' and BauRaum == 'VWAGEN':
                var1, var2 = 'VT09' , 'VT13'
                dict_vts[var1] = vt_merger(dict_vts[var1],dict_vts[var2],[var1, var2]) 
                dict_vts.pop(var2,None)
                dict_vts[var1].to_excel('prueba_VT09_10_13.xlsx',index=False)



            #siempre válido
            if Projekt == 'BH' and BauRaum == 'VWAGEN':
                var1, var2 = 'VTS_06' , 'VTS_07'
                dict_vts[var1] = vt_merger(dict_vts[var1],dict_vts[var2],[var1, var2])
                dict_vts.pop(var2,None)
            
            dict_fk_dict_vts[Sesam_doc + '_' + HandDrive + '_' + BauRaum] = dict_vts 
            
        if Projekt == 'J1':
            for cockpit in dict_fk_dict_vts.keys():
                if 'COCKPIT' in cockpit and 'LL' in cockpit:
                    cockpit_LL = cockpit
                elif 'COCKPIT' in cockpit and 'RL' in cockpit:
                    cockpit_RL = cockpit

            #ojo a esto que está un poco lioso con respecto a lo de arriba... pero it works properly :)
            
            var1, var2 = 'VT02' , 'VT05'
            
            dict_fk_dict_vts[cockpit_LL][var1] = \
             vt_merger(dict_fk_dict_vts[cockpit_LL][var1],
                  dict_fk_dict_vts[cockpit_RL][var1],[var1 + '_LL',var1 + '_RL'])
            dict_fk_dict_vts[cockpit_RL].pop(var1,None)
            
            dict_fk_dict_vts[cockpit_LL][var2] = \
             vt_merger(dict_fk_dict_vts[cockpit_LL][var2],
                  dict_fk_dict_vts[cockpit_RL][var2],[var2 + '_LL',var2 + '_RL'])
            dict_fk_dict_vts[cockpit_RL].pop(var2,None)
        
    return dict_fk_dict_vts


def vt_merger(vt1,vt2,del_var):

    #We will fill in with NON-EXISTING modules in var2 not in var1
    mod2 = vt2['MV intern']
    mod1 = vt1['MV intern']
    vars1 = vt1.columns[15:]
    mod2_miss = ~mod2.isin(mod1)
    #Change the Nummerierung
    #vt1['Nummerierung'] = [str(c) +'_original' for c in vt1['Nummerierung']]
    vt1['Nummerierung'] = [str(c)  for c in vt1['Nummerierung']]
    for idx,mod2_m in enumerate(mod2_miss[mod2_miss].index):
        vt1 = vt1.append(pd.Series(),ignore_index= True)
        vt1.iloc[-1,:15] = vt2.iloc[mod2_m,:15]
        vt1.iloc[-1,15:] = '-'
        vt1 = vt1.sort_index()
        vt1.reset_index(inplace = True, drop = True)
        
    #We will fill in with EXISTING modules in var2 not in var1
    mod2_miss = mod2.isin(mod1)
    #add vars2 in var1
    vars2 = vt2.columns[15:]
    for variant2 in vars2:
        vt1[variant2] = '-'
        mod_var2 = vt2[vt2[variant2].str.upper() == 'X']['MV intern']
        vt1.loc[vt1['MV intern'].isin(mod_var2),variant2] = 'X'
        
    print('VT family '+ del_var[1] + '(now deleted) has been merged with '+ del_var[0])
    
    return vt1

def get_dict_vts_all_FKs(dict_fk_dict_vts):
    ''' 
        Combines the inra & cockpit dictionaries lol with lol and lor with lor.
        We create a random number for each fk, we prevent having two vt from 
        different fk with the same name. Example:
            
            ERROR SITUATION:
                - fk-1: vt01
                -fk-2: vt01
                
            SOLUTION: 
            
               - fk-1: vt01-65 ---> 65 as random number
               - fk-2: vt01-323 ---> 323 as random number
   
    '''
    
    dict_vts_all_FKs = {}
    
    for fk_type,dict_vts in dict_fk_dict_vts.items():
    
        for vt_name, df_vt in dict_vts.items():
            dict_vts_all_FKs[fk_type + '_' + vt_name] = df_vt
# =============================================================================
#             dict_vts_all_FKs[vt_name] = df_vt
# =============================================================================
            
    return dict_vts_all_FKs

def check_variant_diferent_status():
    '''
    Con esta funciónn obtenemos las variantes de dos status diferentes que tiene
    igual nombre pero diferente composición (empleando los nummerierungs)

    Returns
    -------
    None.

    '''
    
    fk1_path = load_fk_file()
    fk2_path = load_fk_file()
    
    fk1_variants = pd.DataFrame()
    fk2_variants = pd.DataFrame()
    
    reporte = []
    
    # Obtenemos un DataFrame con los datos (nummerierungs) de las variantes del 1er FK
    dict_fk1_vts = get_dict_vts(fk1_path)    
    for VT in dict_fk1_vts:
        if not VT.lower().startswith('st'):
            dict_modules = vt.get_dict_variants_nummerierungs_with_VT(dict_fk1_vts[VT], VT)
            df_modules = pd.DataFrame.from_dict(dict_modules, orient = 'index')
            df_modules.insert(0, 'Variants', df_modules.index)
            fk1_variants = pd.concat([fk1_variants, df_modules], ignore_index = True)
        
    # Obtenemos un DataFrame con los datos (nummerierungs) de las variantes del 2er FK
    dict_fk2_vts = get_dict_vts(fk2_path)
    for VT in dict_fk2_vts:
        if not VT.lower().startswith('st'):
            dict_modules = vt.get_dict_variants_nummerierungs_with_VT(dict_fk2_vts[VT], VT)
            df_modules = pd.DataFrame.from_dict(dict_modules, orient = 'index')
            df_modules.insert(0, 'Variants', df_modules.index)
            fk2_variants = pd.concat([fk2_variants, df_modules], ignore_index = True)
    
    # Comparamos los FKs
    for idx1 in fk1_variants.index:
        variant = fk1_variants['Variants'].iloc[idx1]
        pos = fk2_variants.index[(fk2_variants['Variants'].iloc[:] == variant)]
        if len(pos) != 0:
          idx2 = pos[-1]
          #°l1 = len(fk1_variants.loc[idx1])
          #l2 = len(fk2_variants.loc[idx2])
          fk1_aux = fk1_variants.loc[idx1].dropna(axis=1,how='all')
          fk2_aux = fk2_variants.loc[idx2].dropna(axis=1,how='all')
          
          l1 = len(fk1_aux)
          l2 = len(fk2_aux)
          
          # Diferentes cantidades de nummerierungs entonces el contenido es diferente
          if l1 != l2:
              reporte.append(['------------------ different length ----------------------------'])
              reporte.append(fk1_aux.iloc[idx1])
              reporte.append(fk2_aux.iloc[idx2])
          else:
              for i in range(l1):
                  
                  # Diferentes números de nummerierungs
                  if str(fk1_aux.iloc[idx1, i]) != str(fk2_aux.iloc[idx2, i]):
                      reporte.append(['------------------ different content ----------------------------'])
                      reporte.append(fk1_aux.iloc[idx1])
                      reporte.append(fk2_aux.iloc[idx2])
                      break
                  
    df_reporte = pd.DataFrame(reporte)              
    df_reporte.to_excel("report.xlsx", sheet_name = 'report')
    
    return df_reporte

def check_variant_diferent_status(fk1_path, fk2_path):
    '''
    Con esta funciónn obtenemos las variantes de dos status diferentes que tiene
    igual nombre pero diferente composición (empleando los nummerierungs)

    Returns
    -------
    None
    '''

    fk1_variants = pd.DataFrame()
    fk2_variants = pd.DataFrame()
    
    reporte = []
    
    # Obtenemos un DataFrame con los datos (nummerierungs) de las variantes del 1er FK
    dict_fk1_vts = get_dict_vts(fk1_path)    
    for VT in dict_fk1_vts:
        if not VT.lower().startswith('st'):
            dict_modules = vt.get_dict_variants_nummerierungs_with_VT(dict_fk1_vts[VT], VT)
            df_modules = pd.DataFrame.from_dict(dict_modules, orient = 'index')
            df_modules.insert(0, 'Variants', df_modules.index)
            fk1_variants = pd.concat([fk1_variants, df_modules], ignore_index = True)
        
    # Obtenemos un DataFrame con los datos (nummerierungs) de las variantes del 2er FK
    dict_fk2_vts = get_dict_vts(fk2_path)
    for VT in dict_fk2_vts:
        if not VT.lower().startswith('st'):
            dict_modules = vt.get_dict_variants_nummerierungs_with_VT(dict_fk2_vts[VT], VT)
            df_modules = pd.DataFrame.from_dict(dict_modules, orient = 'index')
            df_modules.insert(0, 'Variants', df_modules.index)
            fk2_variants = pd.concat([fk2_variants, df_modules], ignore_index = True)
    
    # Comparamos los FKs
    for idx1 in fk1_variants.index:
        variant = fk1_variants['Variants'].iloc[idx1]
        pos = fk2_variants.index[(fk2_variants['Variants'].iloc[:] == variant)]
        if len(pos) != 0:
          idx2 = pos[-1]
          
          fk1_aux = fk1_variants.loc[idx1].dropna()
          fk2_aux = fk2_variants.loc[idx2].dropna()
          
          l1 = len(fk1_aux)
          l2 = len(fk2_aux)
          
          # Diferentes cantidades de nummerierungs entonces el contenido es diferente
          if l1 != l2:
              reporte.append(['------------------ different length ----------------------------'])
              reporte.append(fk1_aux)
              reporte.append(fk2_aux)
          else:
              for i in range(l1):
                  
                  # Diferentes números de nummerierungs
                  if str(fk1_aux.iloc[i]) != str(fk2_aux.iloc[i]):
                      reporte.append(['------------------ different content ----------------------------'])
                      reporte.append(fk1_aux)
                      reporte.append(fk2_aux)
                      break
                  
    df_reporte = pd.DataFrame(reporte)              
    df_reporte.to_excel("report.xlsx", sheet_name = 'report')
    
    return df_reporte
