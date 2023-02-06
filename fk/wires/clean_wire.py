# -*- coding: utf-8 -*-
"""
Created on Fri Oct  9 10:46:52 2020

@author: eduardo.moreno
"""

import openpyxl

import tkinter.filedialog
import numpy as np
import pandas as pd




def eliminar_columnas_filas():
    
        path = tkinter.filedialog.askopenfilename(filetypes=[("FK", ".xlsx .xls")])
#pedir al usuario el archivo a trabajar
        wb = openpyxl.load_workbook(path)
        hojas = wb.sheetnames
        
        for name in hojas:
            if name.lower().find('wire')!=-1 :
                ws=wb[name]
                
#    ws=wb['VTA']
        
        #ws=wb[name]
#    ws=wb['VTA']
    #Primero recorremos filas basándonos en la columna E (Módulo interno)
        for cell in ws['C']:
            if cell.font.strike:  #el texto está tachado
                ws.delete_rows(cell.row)
        wb.save("Wire_Limpio.xlsx")
        wb.close()



def wirelist_get_family(df_wirelist,familia):
    
    saltar=False
    #familia='226'
    #filtramos la familia del df
    df_wirelist_familia=df_wirelist[df_wirelist['Modulfamilie Zeichnung']==familia]
    
    #dividimos en dos data_frames
    try:
        fila_separacion=np.where(df_wirelist_familia['Ltg-Nr.']=='Ltg-Nr.')[0][0]
    
    except:
        saltar=True
        fila_separacion=1
        
    
    df_1=df_wirelist_familia.iloc[:fila_separacion,:]
    df_2=df_wirelist_familia.iloc[fila_separacion:,:]
    
    
    return df_1,df_2,saltar



import pandas as pd

ruta="C:/Users/eduardo.moreno/Desktop/Ordenador Nuevo Edu/Antenas/D989193-7-F_Konzept_VW276_Inra_LOL_KW22_25_TAB016622_K_ZD20201112_PVS_IC_007.xlsx"
df_2=pd.read_excel(ruta)

def dict_variante_wires(df_2):
    
    #Primera fila a header
    df_2.columns = df_2.iloc[0,:]
    df_2.drop(df_2.index[0],inplace=True)
        
        #Creamos un diccionario para con clave la columna variante y valor lista de wires
        
    dict_variante_wires={}
        
        #prueba=df_2.columns[60:]
        
        #columna='V10'
        #recorremos la lista de columnas y filtramos aquellas que tengan len=2 y empiezen por V
    for columna in df_2.columns:
            #columna='V1'
        if len(str(columna).strip())<=3 and str.lower(str(columna).strip())[0]=='v' and str.lower(str(columna))!='von':
            lista_wires=df_2['Ltg-Nr.'].iloc[:,0][df_2[columna]=='x'].tolist()
            dict_variante_wires[columna]=lista_wires

    
    return dict_variante_wires
    


def df_variante_externalmodul(df_1):
    
    #Eliminamos la primera fila
    df_1.drop(df_1.index[0],inplace=True)
    
    df_variante_externalmodul=df_1[['Ltg-Nr.','Verbindung']]
    
    return df_variante_externalmodul



def dict_externalmodul_wires(dict_variante_wires,df_variante_externalmodul):
    
    dict_externalmodul_wires={}
    
    for key in dict_variante_wires:
        #key='V1'
        df_modulo=df_variante_externalmodul['Verbindung'][df_variante_externalmodul['Ltg-Nr.']==key].tolist()
        
        if len(df_modulo)!=0:
            
            dict_externalmodul_wires[df_modulo[0]]=dict_variante_wires[key]
            
    return dict_externalmodul_wires
    
    
 


   

    
    
    