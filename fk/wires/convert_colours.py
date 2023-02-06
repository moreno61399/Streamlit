# -*- coding: utf-8 -*-
"""
Created on Thu Feb 11 14:31:51 2021

@author: jesus.roldan
"""

from openpyxl.styles import PatternFill, Font
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from tkinter import filedialog
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


import pandas as pd
import numpy as np

def convert_rgb_to_hex(rgb):
    color_string='FF' + ''.join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb])
    return color_string

def convert_colour_to_rgb(colour):

    if colour == 'SW':
        rgb =[0,0,0]
    elif colour == 'GE':
        rgb =[255,255,0]
    elif colour == 'BL':
        rgb =[0,0,255]
    elif colour == 'BR':
        rgb =[165,0,33]
    elif colour == 'RT':
        rgb =[255,0,0]
    elif colour == 'GR':
        rgb =[128,128,128]
    elif colour == 'WS':
        rgb =[255,255,255]
    elif colour == 'VI':
        rgb =[204,0,204]
    elif colour == 'GN':
        rgb =[146,208,80]
    elif colour == 'OR':
        rgb =[255,192,0]
    else:
        print('colour not found')

    return rgb


ruta_FK = filedialog.askopenfilename(title = 'Select F-Konzept',filetypes=[('F-Konzept','.xlsx .xlsm .xls')])

df_fk = pd.read_excel(ruta_FK)
df_fk['Farbe'].fillna(np.nan)

key_column_Farbe = df_fk.columns.get_loc('Farbe')
column_Farbe_1 = key_column_Farbe + 1
column_Farbe_2 = key_column_Farbe + 2

df_fk.insert(column_Farbe_1,'Farbe_1',np.nan)
df_fk.insert(column_Farbe_2,'Farbe_2',np.nan)

wb_fk = Workbook()
ws_wires = wb_fk.active

for r in dataframe_to_rows(df_fk, index=False, header=True):
    ws_wires.append(r)


column_letter =get_column_letter(key_column_Farbe+1)

#analizamos los colores
for cell in ws_wires[column_letter][1:]:
    row = cell.row
    if str(cell.value)!='nan':
        main_colour = cell.value.split('/')[0]
        rgb_main = convert_colour_to_rgb(main_colour)
        hex_main = convert_rgb_to_hex(rgb_main)
        
        if '/' in cell.value:
            secondary_colour = cell.value.split('/')[1]
        else:
            secondary_colour = main_colour
            
        rgb_secondary = convert_colour_to_rgb(secondary_colour)
        hex_secondary = convert_rgb_to_hex(rgb_secondary)
        
        ws_wires.cell(row,column_Farbe_1+1).fill = PatternFill(start_color=hex_main, end_color=hex_main, fill_type = "solid")
        ws_wires.cell(row,column_Farbe_2+1).fill = PatternFill(start_color=hex_secondary, end_color=hex_secondary, fill_type = "solid")

wb_fk.save('base.xlsx')






