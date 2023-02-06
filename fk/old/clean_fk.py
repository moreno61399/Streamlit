
# -*- coding: utf-8 -*-
"""
Spyder Editor
This is a temporary script file.

"""
import openpyxl
import pandas as pd
import tkinter.filedialog
import numpy as np
import os
import sys


path = ''
wb = None
sheets = None


def load_xlsx_file(file_path=None):
    '''
        If we want to use the clean_fk file methods,
        we need to initialize the load of excel file
        by calling first this method.
        
        TODO:
            - Take a deeper look into the scope of global
              variables. It may be a better way to use them.
    '''
    global path
    global sheets
    global wb
                        
    try:
        if file_path is None:
            path = tkinter.filedialog.askopenfilename(filetypes=[("Excel files", ".xlsx .xls .xlsm")])
        else:
            path = file_path
        wb = openpyxl.load_workbook(path, data_only=True)                           # Data only avoids 'formulas'
        sheets = wb.sheetnames
    except Exception as error:
        print('An error ocurred:\n <{}>'.format(error))
        sys.exit()             
                                     
    

def file_size(file_path):
    try:
        size = os.path.getsize(file_path)
        print('Selected file size: {} bytes.'.format(size))
    except:
        print('Unable to acces the file size.')


def red_striked_vt():
    
    '''
        Removes all the rows/columns with red and striked values
    '''
    dict_fk = {}
  
    for name in sheets:
        
        if (name.lower().startswith('vt') or name.lower().startswith('pim') or name.lower().startswith('st')) and len(name) <= 12:
            ws = wb[name]

            #Primero recorremos filas basándonos en la columna E (Módulo interno)
            for cell in ws['E']:
                if cell.font.strike:  #el texto está tachado
                    ws.delete_rows(cell.row)
                #Después recorremos columnas basándonos en la fila 6 (Número de variante)
                for celda in ws[6]:
                    if celda.font.strike:  #el texto está tachado
                        ws.delete_cols(celda.column)
                        
            df = pd.DataFrame(ws.values)   
                                     # From openpyxl woorkbook to pandas DF
            try:
                df = clean_vt(df)
            except Exception as error:
                print('Error cleaning a VT.\n Error: <{}>'.format(error))
                sys.exit()
                
            dict_fk[name] = df.reset_index(drop=True)
            
    return dict_fk



def clean_vt(df):
    ''' 
        Deletes any possible mistake, like random text in wrong cells
        and replaces the variant tokens cells with the corresponding
        v1, v2, v3, v4, ... cell to be able to take all the data as a perfect rectangular table.  
    '''    

    df = df.drop(df.columns[0], axis=1)                                         # Drops the first column(always empty)
    # Deletes all full None columns and we don't get Nontype + str error doing the ''.join... below
    df = df.dropna(axis=1, how='all')                                           
    df_variants_tokens = df.iloc[[4, 5, 6, 7], 15:]                             # A DF with the variant tokens cell,
                                                                                # always starts at col 'Q' and row '5' . . .
    variants_list = []
                                                                                # We concat all the tokens from every variant
    for i in range(len(df_variants_tokens.iloc[0])):    
                                                                                # and we put them on the list.
        reversed_variant_list = df_variants_tokens[df_variants_tokens.columns[i]].tolist()
        
        reversed_variant = ''
        
        for token in reversed_variant_list:
            
            reversed_variant += str(token)
            
        variants_list.append(reverse_variant(reversed_variant))

    df.iloc[8, 15:] = [variant for variant in variants_list]                    # Replace the v1, v2, ... 
    df = df.iloc[8:, 0:]                                                        # with the corresponding variant
    
    df = remove_bad_rows(df)
    
    # Replaces the first row with the columns headers
    new_header = df.iloc[0]                                                     # Grab the firs row for the header
    df = df[1:]                                                                 # take the data less header
    df.columns = new_header                                                     # set the header row as the df header
    
    return df.dropna(how='all')                                                 # Removes the None values
                                                                                # May be better to put this at the begining...

def remove_bad_rows(df):
    ''' 
        Removes the rows with human mistakes
        if the cell value is None or it doesn't have
        the mv_intern format, we replace it with 'eRr0r'.
        Finally we delete all the rows with some eRr0r value.
    '''
    
    f = lambda x: x if x is not None and (len(str(x)) == 13) and x.startswith('1') else 'eRrOr' 
    # Para mayor precisión, añadir más condiciones
    
    df.iloc[1:,3] = df.iloc[1:,3].map(f)
    df = df[df[4] != 'eRrOr']                                                    # Mv intern is col 4 
    
    
    return df


def reverse_variant(reversed_variant):
    ''' 
        From 'C1001LUAA1J13' to '1L13LUAA001C' 
        This is neccesary because when we take the variant tokens,
        we do from top to bottom.
    '''
    try:
        variant = reversed_variant[-4:] + reversed_variant[-8:-4] + reversed_variant[2:5] + reversed_variant[0:2]
    except Exception as error:
        print('Error with a variant format (it may be the variant size).\n Error <{}>'.format(error))
        sys.exit()
        
    return variant


def get_dict_fk():
    ''' Final method, the one that you ask for '''    
    
    dict_fk = red_striked_vt()
    
    return dict_fk






