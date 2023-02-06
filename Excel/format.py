# -*- coding: utf-8 -*-
"""
Created on Thu Aug 26 15:02:43 2021

@author: CPE_Sumitomo
"""
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

import re
import numpy as np

def adjust_column_width(ws, begin, end):
    '''
    Ajusta el ancho de las columna que se encuantran en el rango [begin, end]. 
    Las columnas se numeran comenzando desde uno.

    Parameters
    ----------
    ws : Worksheet 
        Contiene la información de la hoja del Workbook.
    begin : int
        Número de la primera columna a analizar.
    end : int
        Número de la última columna a analizar.

    Returns
    -------
    None.
    '''
    
    i = begin
    for row in ws.iter_cols(begin, end):
        # Variable para guardar la longitud máxima de las celdad de la columna
        column_width = 0
        for cell in row:
            # Para actualizar la máxima longitud de las celdas
            if column_width < len(str(cell.value)):
                column_width = len(str(cell.value))
        # Actualizamos el ancho de la columna       
        ws.column_dimensions[get_column_letter(i)].width = column_width + 1
        # Actualizamos el número de la columna a analizar
        i = i + 1


def adjust_row_height(ws, begin, end, height):
    '''
    Ajusta el alto de las filas que se encuantran en el rango [begin, end] al
    especificado en 'height'. Las filas se numeran comenzando desde uno.

    Parameters
    ----------
    ws : Worksheet 
        Contiene la información de la hoja del Workbook.
    begin : int
        Número de la primera fila a analizar.
    end : int
        Número de la última fila a analizar.
    height: int
        Alto de las columnas.

    Returns
    -------
    None.
    '''
    
    i = begin
    for row in ws.iter_cols(begin, end):
        # Actualizamos el alto de la fila       
        ws.row_dimensions[i].height = height
        # Actualizamos el número de la fila a analizar
        i = i + 1
        
        
def adjust_column_alignement (ws, begin, end, id_horizontal, id_vertical):
    '''
    Ajusta el alignement de las columnas especificadas en el rango [begin, end].

    Parameters
    ----------
    ws : Worksheet 
        Contiene la información de la hoja del Workbook.
    begin : int
        Número de la primera fila a analizar.
    end : int
        Número de la última fila a analizar.
    id_horizontal : int
       Index de la opción para la alineación horizontal. Los valores pueden ser
       los siguientes:
       0 -> general   1 -> left      2 -> center             3 -> right,
       4 -> fill      5 -> justify   6 -> centerContinuous   7 -> distributed
    id_vertical : int
       Index de la opción para la alineación vertical. Los valores pueden ser
       los siguientes:
       0 -> top   1 -> center   2 -> bottom   3 -> justify   4 -> distributed    
    Returns
    -------
    None.
    '''
    horizontal_alignments = ["general", "left", "center", "right", "fill",
                             "justify", "centerContinuous", "distributed"]
    vertical_aligments = ["top", "center", "bottom", "justify", "distributed"]

    for row in ws.iter_cols(begin, end):
        for cell in row:
            cell.alignment = Alignment(horizontal = horizontal_alignments[id_horizontal],
                                       vertical   = vertical_aligments[id_vertical]) 

def make_filtrer(ws):
    '''
    Inserta un autofiltro en la hoja de excel.

    Parameters
    ----------
    ws : Worksheet 
        Contiene la información de la hoja del Workbook.

    Returns
    -------
    None.
    '''
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

def copy_cell_format (cell, new_cell):
    '''
    Copia el formato de 'cell' a 'new_cell'.

    Parameters
    ----------
    cell : Cell
        Celda de la cual copiar el formato.
    new_cell : Cell
        Celda a cambiar el formato de acuerdo al formato de 'cell'.

    Returns
    -------
    None.
    '''

    new_cell.font = copy(cell.font)
    new_cell.fill = copy(cell.fill)
    new_cell.border = copy(cell.border)
    new_cell.alignment = copy(cell.alignment)
    new_cell.protection = copy(cell.protection)
    new_cell.number_format = copy(cell.number_format)
    
    
# workbook = load_workbook(filename="df_formatos1.xlsx")
# adjust_column_width(workbook[workbook.sheetnames[0]], 1, workbook[workbook.sheetnames[0]].max_column)
# adjust_column_alignement(workbook[workbook.sheetnames[0]], 1, workbook[workbook.sheetnames[0]].max_column, 2, 1)
# adjust_row_width(workbook[workbook.sheetnames[0]], 2, 2, 50)
# make_filtrer(workbook[workbook.sheetnames[0]])

# # workbook = adjust_column_width(workbook)
# workbook.save("out.xlsx")
# workbook.close()


def get_colors_before(ws):
    '''
    Función encargada de obtener los colores de una lista de hojas excel (solo 
    tiene en cuenta los colores rojo y gris). Para ello recorre cada una de las
    celdas si el color de la fuente es rojo adiciona al valor de la celda el string 
    '_red' y si el color de la fuente es gris adiciona al valor de la celda el 
    string '_grey'.

    Parameters
    ----------
    wb : Workbook
        Contiene la información del excel.
    lista_names : lista de String
        Lista de los nombres de las hojas del excel a analizar.

    Returns
    -------
    wb : Workbook
        Contiene la información del excel con los colores añadidos.

    '''    
    try:
        
           RED    = 'FFFF0000' 
           YELLOW = 'FFFFFF00'
           BLUE   = 'FF0000FF'
           GREEN  = 'FF00FF00'
           #GREY   = (166,166,166)
    
           for row in ws.iter_rows(ws.min_row, ws.max_row):
               for cell in row:
                   # Chequeamos que la celda tenga color y no esté vacía
                   if cell.font.color == None:
                      continue
                   
                   if cell.value == None:
                      cell.value = '' 
                       
                   if len(str(cell.font.color.rgb)) == 8:       # Hex value (RRGGBB)
                      
                       color = str(cell.font.color.rgb)
                       
                       if color == RED:
                          cell.value = str(cell.value) + '_red'
                       
                       if str(cell.fill.fgColor.rgb) == YELLOW:
                          cell.value = str(cell.value) + '_yellow' 
                          
                       if str(cell.fill.fgColor.rgb) == GREEN:
                          cell.value = str(cell.value) + '_green'    
                       
                       if color == BLUE:
                          cell.value = str(cell.value) + '_blue'
                       
                       # Color gris
                       x = str(cell.font.color.rgb)
                       if x[2:4] == x[4:6] and x[4:6] == x[6:]:
                          cell.value = str(cell.value) + '_grey'  
                   else:
                       if cell.font.color.theme == 6 or (cell.font.color.tint != 0.0 and \
                          (cell.font.color.theme == 0 or cell.font.color.theme == 1)):
                          cell.value =  str(cell.value) + '_grey'
    except:
       print("El valor ha de ser 'BOM' o 'WIRE'")
    
    return ws

def set_colors_after(ws):
    '''
    Función encargada de poner los colores en una hoja de excel. Recorre cada 
    una de las celdas de la hoja, si la celda contiene el string '_yellow' pone 
    el background de la celda en amarrillo, si contiene '_blue' cambia el color 
    de fuente a azul, si contiene el '_red' cambia el color de fuente a rojo y si
    si contiene el '_grey' cambia el color de fuente a gris.

    Parameters
    ----------
    wb : Workbook
        Contiene la información del excel.
    sheet_name : String
        Nombre de la hoja del excel a analizar.

    Returns
    -------
    None.
    '''
    
    # Color de background amarillo    
    YELLOWFILL = PatternFill(start_color = 'FFFF00', end_color = 'FFFF00', \
                             fill_type   = 'solid')
    # Color de background verde    
    GREENFILL = PatternFill(start_color = '00FF00', end_color = '00FF00', \
                             fill_type   = 'solid')
    
    # Color de background azul claro    
    LIGHTBLUEFILL = PatternFill(start_color = 'CCFFFF', end_color = 'CCFFFF', \
                             fill_type   = 'solid')
    
    # Color de background naranja  
    ORANGEFILL = PatternFill(start_color = 'FF9900', end_color = 'FF9900', \
                             fill_type   = 'solid')
        
    
    # Recorremos todas las celdas
    for row in ws.iter_rows(ws.min_row, ws.max_row):
        con = 0
        for cell in row:
            
            # Poner color de font azul            
            if '_blue' in str(cell.value):
                cell.font  = Font(color = '0000FF')
                cell.value = str(cell.value).replace('_blue', '')
                
            # Poner color de font rojo y strike
            if '_red'  in str(cell.value):
                cell.font  = Font(color = 'FF0000', strike = True)
                cell.value = str(cell.value).replace('_red', '')
            
            # Poner color de font rojo sin strike
            if '_rreedd'  in str(cell.value):
                cell.font  = Font(color = 'FF0000', strike = False)
                cell.value = str(cell.value).replace('_rreedd', '')
                
            # Poner color de font gris
            if '_grey' in str(cell.value):
                cell.font  = Font(color = 'A6A6A6')
                cell.value = str(cell.value).replace('_grey', '')
                
            # Poner color de font azul            
            if '_lightblue' in str(cell.value):
                cell.fill  =  LIGHTBLUEFILL
                cell.value = str(cell.value).replace('_lightblue', '')
                
            # Poner color de background amarillo
            if '_yellow' in str(cell.value):
                cell.fill  = YELLOWFILL
                cell.value = str(cell.value).replace('_yellow', '')
            
            # Poner color de background naranja
            if '_orange' in str(cell.value):
                cell.fill  = ORANGEFILL
                cell.value = str(cell.value).replace('_orange', '')

            
            # Color verde de background, indica que se debe verificar manualmente su información
            if '_green' in str(cell.value):
                cell.fill  =  GREENFILL
                cell.value = str(cell.value).replace('_green', '')
                

            
            if con > 0 and str(cell.value).isnumeric() == True:
               cell.value = int(cell.value) 
            
            num_format = re.compile("[0-9]*\.[0-9]+$")
            isfloat = re.match(num_format, str(cell.value))
            if isfloat:
               cell.value = str(cell.value).replace('.', ',')
            
            con = con + 1

def set_cell_color(cell, color):
    '''
    Poner un color espesífico a una celda.

    Parameters
    ----------
    cell : openpyxl.cell
        Celda a la cual se le cambiará el color.
    color : String
        Color en formato hexadecimal.

    Returns
    -------
    None.
    '''
    
    # Color de background   
    COLOR = PatternFill(start_color = color, end_color = color, \
                        fill_type   = 'solid')
    cell.fill  =  COLOR     
        
'''Data taken from internet for working with colours rgb,hex, theme and tints'''


from colorsys import rgb_to_hls, hls_to_rgb
# From: https://stackoverflow.com/questions/58429823/getting-excel-cell-background-themed-color-as-hex-with-openpyxl/58443509#58443509
#   which refers to: https://pastebin.com/B2nGEGX2 (October 2020)
#       Updated to use list(elem) instead of the deprecated elem.getchildren() method
#       which has now been removed completely from Python 3.9 onwards.
#

#https://bitbucket.org/openpyxl/openpyxl/issues/987/add-utility-functions-for-colors-to-help

RGBMAX = 0xff  # Corresponds to 255
HLSMAX = 240  # MS excel's tint function expects that HLS is base 240. see:
# https://social.msdn.microsoft.com/Forums/en-US/e9d8c136-6d62-4098-9b1b-dac786149f43/excel-color-tint-algorithm-incorrect?forum=os_binaryfile#d3c2ac95-52e0-476b-86f1-e2a697f24969

def rgb_to_ms_hls(red, green=None, blue=None):
    """Converts rgb values in range (0,1) or a hex string of the form '[#aa]rrggbb' to HLSMAX based HLS, (alpha values are ignored)"""
    if green is None:
        if isinstance(red, str):
            if len(red) > 6:
                red = red[-6:]  # Ignore preceding '#' and alpha values
            blue = int(red[4:], 16) / RGBMAX
            green = int(red[2:4], 16) / RGBMAX
            red = int(red[0:2], 16) / RGBMAX
        else:
            red, green, blue = red
    h, l, s = rgb_to_hls(red, green, blue)
    return (int(round(h * HLSMAX)), int(round(l * HLSMAX)), int(round(s * HLSMAX)))

def ms_hls_to_rgb(hue, lightness=None, saturation=None):
    """Converts HLSMAX based HLS values to rgb values in the range (0,1)"""
    if lightness is None:
        hue, lightness, saturation = hue
    return hls_to_rgb(hue / HLSMAX, lightness / HLSMAX, saturation / HLSMAX)

def rgb_to_hex(red, green=None, blue=None):
    """Converts (0,1) based RGB values to a hex string 'rrggbb'"""
    if green is None:
        red, green, blue = red
    return ('%02x%02x%02x' % (int(round(red * RGBMAX)), int(round(green * RGBMAX)), int(round(blue * RGBMAX)))).upper()


def get_theme_colors(wb):
    """Gets theme colors from the workbook"""
    # see: https://groups.google.com/forum/#!topic/openpyxl-users/I0k3TfqNLrc
    from openpyxl.xml.functions import QName, fromstring
    xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    root = fromstring(wb.loaded_theme)
    themeEl = root.find(QName(xlmns, 'themeElements').text)
    colorSchemes = themeEl.findall(QName(xlmns, 'clrScheme').text)
    firstColorScheme = colorSchemes[0]

    colors = []

    for c in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6']:
        accent = firstColorScheme.find(QName(xlmns, c).text)
        for i in list(accent): # walk all child nodes, rather than assuming [0]
            if 'window' in i.attrib['val']:
                colors.append(i.attrib['lastClr'])
            else:
                colors.append(i.attrib['val'])

    return colors

def tint_luminance(tint, lum):
    """Tints a HLSMAX based luminance"""
    # See: http://ciintelligence.blogspot.co.uk/2012/02/converting-excel-theme-color-and-tint.html
    if tint < 0:
        return int(round(lum * (1.0 + tint)))
    else:
        return int(round(lum * (1.0 - tint) + (HLSMAX - HLSMAX * (1.0 - tint))))

def theme_and_tint_to_rgb(wb, theme, tint):
    """Given a workbook, a theme number and a tint return a hex based rgb"""
    rgb = get_theme_colors(wb)[theme]
    h, l, s = rgb_to_ms_hls(rgb)
    return rgb_to_hex(ms_hls_to_rgb(h, tint_luminance(tint, l), s))



